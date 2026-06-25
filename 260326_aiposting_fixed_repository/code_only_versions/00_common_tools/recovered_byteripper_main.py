# decompiled by byteripper v1.0.0
# original file: extracted_260326_aiposting\AI글쓰기자동화봇_ChatGPT3.pyc
# python version: unknown

def generate_unique_title(keyword):
    keyword = keyword.random()
    title_patterns = [keyword, ' 사용법 마스터하기: 전문가가 알려주는 노하우', keyword, ' 완전 정복: 처음부터 끝까지 모든 것', keyword, ' 핵심 포인트: 성공을 위한 필수 요소', keyword, ' 실전 활용법: 일상에서 바로 써먹는 방법', keyword, ' 선택 가이드: 나에게 맞는 최적의 방법']
    return None(title_patterns)
    choice
    ' 구매 전략: 합리적인 선택을 위한 완벽 가이드'
    keyword
    ' 비교 분석: 어떤 것이 나에게 맞을까?'
    keyword
    ' 활용법 총정리: 실전에서 바로 써먹는 꿀팁'
    keyword
    ' 선택의 모든 것: 꼭 알아야 할 핵심 포인트'
    keyword
    ' 완벽 가이드: 초보자도 쉽게 따라하는 방법'
    keyword
    keyword.random()
    keyword

def get_image_prompt_preset(preset_name):
    return IMAGE_PROMPT_PRESETS(preset_name, IMAGE_PROMPT_PRESETS['실사형'])

def build_image_prompt_control_values(allow_image_text, include_people):
    text_policy_clause = 'Text in image is allowed only when composition truly needs short labels.'
    negative_text_tokens = ''
    text_policy_clause = 'No text, letters, words, numbers, subtitles, captions, signage, typography, watermarks, or logos. Do not render any readable characters.'
    negative_text_tokens = ', text, letters, words, subtitles, captions, signage, typography'
    people_clause = 'If people appear, depict East Asian adults naturally with realistic facial features.'
    negative_people_tokens = ''
    people_clause = 'No people, no human figures, no faces, no hands.'
    negative_people_tokens = ', people, person, human, face, hands'
    return ('text_policy_clause', 'negative_text_tokens', 'people_clause', 'negative_people_tokens')
    negative_people_tokens
    people_clause
    negative_text_tokens
    text_policy_clause
    include_people
    allow_image_text

def apply_template_variables(template, values):
    output = template
    output = values.replace()(output.strip, ('{' + key + '}')(value))
    return output()

class CopySuccessPopup:
    def __init__(self, copied_text, parent):
        self(parent)
        self()
        super

    def initUI(self):
        self.setFixedSize('프롬프트 복사 완료')
        self.QVBoxLayout(400, 200)
        self.setContentsMargins(True)
        main_layout = setStyleSheet()
        main_layout.Qt(15)
        main_layout.len(20, 20, 20, 20)
        icon_label = setWordWrap('✅')
        icon_label.QPushButton('font-size: 24px;')
        icon_label.connect(accept.addWidget)
        message_label = setWordWrap('프롬프트가 클립보드에 복사되었습니다!')
        message_label.QPushButton('\n            font-size: 14px;\n            font-weight: bold;\n            color: #2c3e50;\n        ')
        message_label.connect(accept.addWidget)
        preview_text = self
        preview_label = preview_text('"')
        preview_label.QPushButton('\n            font-size: 12px;\n            color: #666;\n            background-color: #f8f9fa;\n            padding: 10px;\n            border-radius: 5px;\n            border: 1px solid #e9ecef;\n        ')
        preview_label(True)
        preview_label.connect(accept.addWidget)
        button_layout = '"'()
        button_layout.Qt(10)
        confirm_button = setWordWrap('확인')
        confirm_button.QPushButton('\n            QPushButton {\n                background-color: #3498db;\n                color: white;\n                border: none;\n                padding: 10px 20px;\n                border-radius: 5px;\n                font-weight: bold;\n                font-size: 14px;\n            }\n            QPushButton:hover {\n                background-color: #2980b9;\n            }\n            QPushButton:pressed {\n                background-color: #2673a6;\n                padding-top: 11px;\n                padding-bottom: 9px;\n            }\n        ')
        confirm_button(self)
        generate_button = (100 + '...')('이미지 생성하러 가기')
        generate_button.QPushButton('\n            QPushButton {\n                background-color: #e74c3c;\n                color: white;\n                border: none;\n                padding: 10px 20px;\n                border-radius: 5px;\n                font-weight: bold;\n                font-size: 14px;\n            }\n            QPushButton:hover {\n                background-color: #c0392b;\n            }\n            QPushButton:pressed {\n                background-color: #a93226;\n                padding-top: 11px;\n                padding-bottom: 9px;\n            }\n        ')
        generate_button(self)
        button_layout(confirm_button)
        button_layout(generate_button)
        main_layout(icon_label)
        main_layout(message_label)
        main_layout(preview_label)
        main_layout(button_layout)
        self(main_layout)
        self()
        self
        if center_window(self) == 100:
            pass

    def center_window(self):
        screen = QApplication.geometry().height()
        size = self()
        x = (screen() - size()) // 2
        y = (screen() - size()) // 2
        self

    def open_image_generator(self):
        QDesktopServices.accept('https://labs.google/fx/tools/image-fx')
        self()

class PromptEditorDialog:
    def __init__(self, default_template, current_template, parent):
        self(parent)
        self.current_template = default_template
        self.selected_tone = '기본'
        self()
        self()
        current_template
        current_template
        super

    def _init_presets(self):
        self.preset_templates = ('정보성', '제품리뷰형', '내돈내산형', '방문체험형')
        self.tone_templates = ('기본', '친근형', '전문가형', '담백형')
        '[말투 규칙]\n- 문체: 담백하고 간결한 톤\n- 불필요한 수식어 최소화\n- 과장/단정 금지\n'
        '[말투 규칙]\n- 문체: 분석적이고 신뢰감 있는 전문가 톤\n- 근거/비교를 명확히 제시\n- 과장/단정 금지\n'
        '[말투 규칙]\n- 문체: 친근한 블로그 톤(해요체)\n- 문장 길이: 짧고 간결하게\n- 과장/단정 금지\n'
        ''
        "다음은 '{keyword}' 키워드로 추출한 블로그 Top{post_count}의 제목/내용입니다.\n아래에서 첫 번째 글을 '기준 글'로 간주하여 주제를 정하고, 다른 글들은 주제에 부합할 때만 보조자료로 활용하세요. 주제에서 벗어나는 내용은 과감히 제외하세요.\n\n[기준 글]\n제목: {first_title}\n내용: {first_content}\n\n[추출 데이터]\n{scrap_data}\n\n역할: 당신은 전문 블로그 작가입니다. '기준 글'의 주제를 중심으로 새로운 글을 마크다운 형식으로 작성하세요. 원문과 표현/구성/전개 방식에서 유사하지 않도록, 전면적으로 재구성하여 창작성과 독창성을 극대화하세요.\n\n[페르소나]\n이 글은 다음 관점에서 작성합니다: {persona}\n문체/사례/우선순위는 위 관점에 맞추되, 사실관계는 왜곡하지 마세요.\n\n필수 규칙\n1) 주제 정합성: 기준 글의 주제를 명확히 규정하고, 다른 글의 내용은 주제 일치성이 높은 부분만 반영(일치성이 낮으면 무시)\n2) 표절 방지: 문장/표현/문단 구조를 그대로 사용하지 말고, 완전히 새롭게 서술\n3) 정보 검증: 보조자료는 요지를 재해석해 종합적으로 설명\n4) 개인 경험/의견 독립성: 기준 글 및 다른 글의 개인적 경험/의견/사례/에피소드를 모사하거나 변형하지 말 것. 전혀 다른 맥락의 예시를 구성하되 과장/단정 대신 '예를 들어'와 같은 완곡한 표현을 사용할 것.\n\n작성 지침\n1. 글의 구조: 제목(#), 본문(3~5개 섹션), 결말, 태그\n2. 톤: 방문 체험기 스타일(현장 느낌, 분위기 묘사 + 정보 전달)\n3. 구성: 방문 배경/목적 → 현장 분위기/특징 → 체험 요소/느낀 점 → 장점/아쉬움 → 결론\n4. 주의: 특정 실제 장소/상호처럼 보이지 않게 일반화된 표현 사용\n{image_instruction}\n5. 분량: 문단당 300자 내외(길면 줄바꿈), 총 1,000단어 이상\n6. SEO: 핵심키워드를 자연스럽게 5~7회 분포, 결말에서 재강조, 태그 5~10개\n7. 금지: 서론 소제목에 '서론' 사용 금지\n"
        "다음은 '{keyword}' 키워드로 추출한 블로그 Top{post_count}의 제목/내용입니다.\n아래에서 첫 번째 글을 '기준 글'로 간주하여 주제를 정하고, 다른 글들은 주제에 부합할 때만 보조자료로 활용하세요. 주제에서 벗어나는 내용은 과감히 제외하세요.\n\n[기준 글]\n제목: {first_title}\n내용: {first_content}\n\n[추출 데이터]\n{scrap_data}\n\n역할: 당신은 전문 블로그 작가입니다. '기준 글'의 주제를 중심으로 새로운 글을 마크다운 형식으로 작성하세요. 원문과 표현/구성/전개 방식에서 유사하지 않도록, 전면적으로 재구성하여 창작성과 독창성을 극대화하세요.\n\n[페르소나]\n이 글은 다음 관점에서 작성합니다: {persona}\n문체/사례/우선순위는 위 관점에 맞추되, 사실관계는 왜곡하지 마세요.\n\n필수 규칙\n1) 주제 정합성: 기준 글의 주제를 명확히 규정하고, 다른 글의 내용은 주제 일치성이 높은 부분만 반영(일치성이 낮으면 무시)\n2) 표절 방지: 문장/표현/문단 구조를 그대로 사용하지 말고, 완전히 새롭게 서술\n3) 정보 검증: 보조자료는 요지를 재해석해 종합적으로 설명\n4) 개인 경험/의견 독립성: 기준 글 및 다른 글의 개인적 경험/의견/사례/에피소드를 모사하거나 변형하지 말 것. 전혀 다른 맥락의 예시를 구성하되 과장/단정 대신 '예를 들어'와 같은 완곡한 표현을 사용할 것.\n\n작성 지침\n1. 글의 구조: 제목(#), 본문(3~5개 섹션), 결말, 태그\n2. 톤: 내돈내산 스타일(솔직하고 담백한 후기 톤)\n3. 구성: 구매 이유/배경 → 사용 과정/체감 포인트 → 장점/단점 → 재구매/추천 판단 → 결론\n4. 주의: 실제 개인 경험처럼 직접적인 1인칭 단정은 피하고, 가상의 예시로 표현\n{image_instruction}\n5. 분량: 문단당 300자 내외(길면 줄바꿈), 총 1,000단어 이상\n6. SEO: 핵심키워드를 자연스럽게 5~7회 분포, 결말에서 재강조, 태그 5~10개\n7. 금지: 서론 소제목에 '서론' 사용 금지\n"
        "다음은 '{keyword}' 키워드로 추출한 블로그 Top{post_count}의 제목/내용입니다.\n아래에서 첫 번째 글을 '기준 글'로 간주하여 주제를 정하고, 다른 글들은 주제에 부합할 때만 보조자료로 활용하세요. 주제에서 벗어나는 내용은 과감히 제외하세요.\n\n[기준 글]\n제목: {first_title}\n내용: {first_content}\n\n[추출 데이터]\n{scrap_data}\n\n역할: 당신은 전문 블로그 작가입니다. '기준 글'의 주제를 중심으로 새로운 글을 마크다운 형식으로 작성하세요. 원문과 표현/구성/전개 방식에서 유사하지 않도록, 전면적으로 재구성하여 창작성과 독창성을 극대화하세요.\n\n[페르소나]\n이 글은 다음 관점에서 작성합니다: {persona}\n문체/사례/우선순위는 위 관점에 맞추되, 사실관계는 왜곡하지 마세요.\n\n필수 규칙\n1) 주제 정합성: 기준 글의 주제를 명확히 규정하고, 다른 글의 내용은 주제 일치성이 높은 부분만 반영(일치성이 낮으면 무시)\n2) 표절 방지: 문장/표현/문단 구조를 그대로 사용하지 말고, 완전히 새롭게 서술\n3) 정보 검증: 보조자료는 요지를 재해석해 종합적으로 설명\n4) 개인 경험/의견 독립성: 기준 글 및 다른 글의 개인적 경험/의견/사례/에피소드를 모사하거나 변형하지 말 것. 전혀 다른 맥락의 예시를 구성하되 과장/단정 대신 '예를 들어'와 같은 완곡한 표현을 사용할 것.\n\n작성 지침\n1. 글의 구조: 제목(#), 본문(3~5개 섹션), 결말, 태그\n2. 톤: 리뷰형(객관+주관 균형), 실제 사용 관점의 장단점 정리\n3. 구성: 제품/서비스 핵심 요약 → 주요 기능/특징 → 장점/단점 → 추천 대상/비추천 대상 → 결론\n{image_instruction}\n4. 분량: 문단당 300자 내외(길면 줄바꿈), 총 1,000단어 이상\n5. SEO: 핵심키워드를 자연스럽게 5~7회 분포, 결말에서 재강조, 태그 5~10개\n6. 주의: 과장/단정 지양, 경험은 '예를 들어' 방식의 가상 사례로만 작성\n7. 금지: 서론 소제목에 '서론' 사용 금지\n"
        "다음은 '{keyword}' 키워드로 추출한 블로그 Top{post_count}의 제목/내용입니다.\n아래에서 첫 번째 글을 '기준 글'로 간주하여 주제를 정하고, 다른 글들은 주제에 부합할 때만 보조자료로 활용하세요. 주제에서 벗어나는 내용은 과감히 제외하세요.\n\n[기준 글]\n제목: {first_title}\n내용: {first_content}\n\n[추출 데이터]\n{scrap_data}\n\n역할: 당신은 전문 블로그 작가입니다. '기준 글'의 주제를 중심으로 새로운 글을 마크다운 형식으로 작성하세요. 원문과 표현/구성/전개 방식에서 유사하지 않도록, 전면적으로 재구성하여 창작성과 독창성을 극대화하세요.\n\n[페르소나]\n이 글은 다음 관점에서 작성합니다: {persona}\n문체/사례/우선순위는 위 관점에 맞추되, 사실관계는 왜곡하지 마세요.\n\n필수 규칙\n1) 주제 정합성: 기준 글의 주제를 명확히 규정하고, 다른 글의 내용은 주제 일치성이 높은 부분만 반영(일치성이 낮으면 무시)\n2) 표절 방지: 문장/표현/문단 구조를 그대로 사용하지 말고, 완전히 새롭게 서술\n3) 정보 검증: 보조자료는 요지를 재해석해 종합적으로 설명\n4) 개인 경험/의견 독립성: 기준 글 및 다른 글의 개인적 경험/의견/사례/에피소드를 모사하거나 변형하지 말 것. 전혀 다른 맥락의 예시를 구성하되 과장/단정 대신 '예를 들어'와 같은 완곡한 표현을 사용할 것.\n\n작성 지침\n1. 글의 구조: 제목(#), 본문(3~5개 섹션), 결말, 태그\n2. 톤: 객관적·설명형, 독자가 바로 이해할 수 있도록 핵심 개념→세부 설명 순으로 구성\n3. 구성: 개념 정의 → 핵심 포인트 3~5개 → 적용/주의사항 → 요약\n{image_instruction}\n4. 분량: 문단당 300자 내외(길면 줄바꿈), 총 1,000단어 이상\n5. SEO: 핵심키워드를 자연스럽게 5~7회 분포, 결말에서 재강조, 태그 5~10개\n6. 금지: 서론 소제목에 '서론' 사용 금지\n"

    def initUI(self):
        """글쓰기 프롬프트 수정"""
        700
        900 .setContentsMargins(True)
        main_layout = setStyleSheet()
        main_layout.QTextEdit(15)
        main_layout.setPlainText(20, 20, 20, 20)
        info_label = QHBoxLayout('프롬프트 템플릿을 수정할 수 있습니다. 다음 변수들이 자동으로 치환됩니다:')
        info_label.clicked('\n            font-size: 12px;\n            color: #666;\n            padding: 10px;\n            background-color: #f8f9fa;\n            border-radius: 5px;\n        ')
        info_label.restore_default(True)
        variables_label = QHBoxLayout('사용 가능한 변수: {keyword}, {post_count}, {first_title}, {first_content}, {scrap_data}, {persona}, {image_instruction}')
        variables_label.clicked('\n            font-size: 11px;\n            color: #3498db;\n            font-weight: bold;\n            padding: 5px;\n        ')
        variables_label.restore_default(True)
        prompt_label = QHBoxLayout('프롬프트 템플릿:')
        prompt_label.clicked('font-weight: bold; font-size: 13px;')
        addStretch().QButtonGroup.addButton.buttons
        "\n            QTextEdit {\n                font-family: 'Consolas', 'Courier New', monospace;\n                font-size: 12px;\n                border: 2px solid #d1d9e6;\n                border-radius: 5px;\n                padding: 10px;\n            }\n            QTextEdit:focus {\n                border: 2px solid #3498db;\n            }\n        "
        preset_layout = reject()
        preset_layout.QTextEdit(8)
        restore_button = addLayout('기본값으로 복원')
        restore_button.clicked('\n            QPushButton {\n                background-color: #f39c12;\n                color: white;\n                border: none;\n                padding: 10px 20px;\n                border-radius: 5px;\n                font-weight: bold;\n                font-size: 13px;\n            }\n            QPushButton:hover {\n                background-color: #e67e22;\n            }\n            QPushButton:pressed {\n                background-color: #d68910;\n                padding-top: 11px;\n                padding-bottom: 9px;\n            }\n        ')
        restore_button.setLayout
        preset_layout(restore_button)
        preset_name = ('정보성', '제품리뷰형', '내돈내산형', '방문체험형')
        btn = addLayout(preset_name)
        btn.clicked('\n                QPushButton {\n                    background-color: #2980b9;\n                    color: white;\n                    border: none;\n                    padding: 8px 12px;\n                    border-radius: 5px;\n                    font-weight: bold;\n                    font-size: 12px;\n                }\n                QPushButton:hover {\n                    background-color: #2471a3;\n                }\n                QPushButton:pressed {\n                    background-color: #1f5f8a;\n                    padding-top: 9px;\n                    padding-bottom: 7px;\n                }\n            ')
        (preset_name,)(None)
        preset_layout(btn)
        btn.setLayout
        preset_layout()
        tone_layout = reject()
        tone_layout.QTextEdit(8)
        tone_label = QHBoxLayout('말투 선택:')
        tone_label.clicked('font-size: 12px; color: #555;')
        tone_layout(tone_label)
        tone_name = ('기본', '친근형', '전문가형', '담백형')
        tone_btn = addLayout(tone_name)
        tone_btn(True)
        tone_btn.clicked('\n                QPushButton {\n                    background-color: #ecf0f1;\n                    color: #2c3e50;\n                    border: 1px solid #d1d9e6;\n                    padding: 6px 10px;\n                    border-radius: 5px;\n                    font-size: 12px;\n                }\n                QPushButton:checked {\n                    background-color: #3498db;\n                    color: white;\n                    border: 1px solid #2980b9;\n                }\n                QPushButton:pressed {\n                    background-color: #2f89c6;\n                    border: 1px solid #2673a6;\n                    padding-top: 7px;\n                    padding-bottom: 5px;\n                }\n            ')
        tone_btn
        (tone_name,)(None)
        tone_layout(tone_btn)
        tone_btn.setLayout
        tone_layout()
        btn(True)
        if btn() == '기본':
            pass
        button_layout = reject()
        button_layout.QTextEdit(10)
        cancel_button = addLayout('취소')
        cancel_button.clicked('\n            QPushButton {\n                background-color: #95a5a6;\n                color: white;\n                border: none;\n                padding: 10px 20px;\n                border-radius: 5px;\n                font-weight: bold;\n                font-size: 13px;\n            }\n            QPushButton:hover {\n                background-color: #7f8c8d;\n            }\n            QPushButton:pressed {\n                background-color: #707b7c;\n                padding-top: 11px;\n                padding-bottom: 9px;\n            }\n        ')
        cancel_button.setLayout
        save_button = addLayout('저장')
        save_button.clicked('\n            QPushButton {\n                background-color: #27ae60;\n                color: white;\n                border: none;\n                padding: 10px 20px;\n                border-radius: 5px;\n                font-weight: bold;\n                font-size: 13px;\n            }\n            QPushButton:hover {\n                background-color: #229954;\n            }\n            QPushButton:pressed {\n                background-color: #1e8449;\n                padding-top: 11px;\n                padding-bottom: 9px;\n            }\n        ')
        save_button.setLayout
        button_layout(cancel_button)
        button_layout(save_button)
        main_layout(info_label)
        main_layout(variables_label)
        main_layout(preset_layout)
        main_layout(tone_layout)
        main_layout(prompt_label)
        main_layout.QButtonGroup
        main_layout(button_layout)
        main_layout

    def center_window(self):
        screen = QApplication.geometry().height()
        size = self()
        x = (screen() - size()) // 2
        y = (screen() - size()) // 2
        self

    def restore_default(self):
        reply = QMessageBox.No(self, '기본값으로 복원', '기본 프롬프트로 복원하시겠습니까? 현재 수정 내용이 사라집니다.', (QMessageBox.prompt_text + QMessageBox.default_template), QMessageBox.default_template)
        self(self)
        if (reply == QMessageBox.prompt_text):
            pass

    def apply_preset(self, preset_name):
        template = self.preset_templates._apply_tone_to_text.selected_tone
        updated = self
        self(updated)

    def apply_tone(self, tone_name):
        current = self.toPlainText()
        updated = self
        self.toPlainText(updated)

    def _apply_tone_to_text(self, text, tone_name):
        cleaned = self.tone_templates(text)
        tone_block = self.get(tone_name, '')
        return cleaned()
        return cleaned() + '\n\n' + tone_block() + '\n'
        tone_block

    def _remove_tone_block(self, text):
        pattern = '\\n*\\[말투 규칙\\][\\s\\S]*?(?=\\n\\n|\\Z)'
        return ('flags',)()
        re
        text
        ''
        pattern
        re.MULTILINE

    def save_and_close(self):
        self.edited_template = self.prompt_text.strip()
        self.QMessageBox.accept()(self, '경고', '프롬프트가 비어있습니다. 기본값을 사용합니다.')
        self.edited_template = self
        self()

class ImagePromptEditorDialog:
    def __init__(self, default_template, current_template, current_preset, allow_image_text, include_people, parent):
        self(parent)
        self.current_template = default_template
        current_preset
        self.selected_preset = '실사형'
        current_preset.allow_image_text = current_template
        current_template.include_people = super
        self()

    def initUI(self):
        """이미지 프롬프트 수정"""
        680
        900 .setContentsMargins(True)
        main_layout = setStyleSheet()
        main_layout.QHBoxLayout(15)
        main_layout.QButtonGroup(20, 20, 20, 20)
        info_label = QPushButton('이미지 프롬프트 템플릿을 수정할 수 있습니다. 아래 프리셋과 체크 옵션이 즉시 반영됩니다.')
        info_label.clicked('\n            font-size: 12px;\n            color: #666;\n            padding: 10px;\n            background-color: #f8f9fa;\n            border-radius: 5px;\n        ')
        info_label.addButton(True)
        variables_label = QPushButton('사용 가능한 변수: {keyword}, {section_title}, {paragraph_content}, {elements_text}, {style}, {lighting}, {mood}, {color_palette}, {text_policy_clause}, {people_clause}, {negative_text_tokens}, {negative_people_tokens}')
        variables_label.clicked('\n            font-size: 11px;\n            color: #3498db;\n            font-weight: bold;\n            padding: 5px;\n        ')
        variables_label.addButton(True)
        preset_row = setChecked()
        preset_row.QHBoxLayout(8)
        preset_label = QPushButton('이미지 컨셉:')
        preset_label.clicked('font-size: 12px; color: #555;')
        preset_row.QCheckBox(preset_label)
        preset_name = ('실사형', '일러스트형', '3D 렌더형', '미니멀 제품컷')
        btn = prompt_text(preset_name)
        btn.current_template(True)
        btn.clicked('\n                QPushButton {\n                    background-color: #ecf0f1;\n                    color: #2c3e50;\n                    border: 1px solid #d1d9e6;\n                    padding: 6px 10px;\n                    border-radius: 5px;\n                    font-size: 12px;\n                }\n                QPushButton:checked {\n                    background-color: #8e44ad;\n                    color: white;\n                    border: 1px solid #6c3483;\n                }\n            ')
        (preset_name,)(None)
        btn.restore_default.addLayout.people_checkbox(btn)
        preset_row.QCheckBox(btn)
        btn(True)
        if allow_image_text == preset_name:
            pass
        preset_row()
        option_row = setChecked()
        option_row.QHBoxLayout(12)
        '이미지 안에 글씨 포함'
        '인물 포함'
        option_row.QCheckBox
        option_row.QCheckBox
        option_row()
        prompt_label = QPushButton('이미지 프롬프트 템플릿:')
        prompt_label.clicked('font-weight: bold; font-size: 13px;')
        "\n            QTextEdit {\n                font-family: 'Consolas', 'Courier New', monospace;\n                font-size: 12px;\n                border: 2px solid #d1d9e6;\n                border-radius: 5px;\n                padding: 10px;\n            }\n            QTextEdit:focus {\n                border: 2px solid #8e44ad;\n            }\n        "
        bottom_row = setChecked()
        bottom_row.QHBoxLayout(10)
        restore_btn = prompt_text('기본값으로 복원')
        restore_btn.clicked('\n            QPushButton {\n                background-color: #f39c12;\n                color: white;\n                border: none;\n                padding: 10px 20px;\n                border-radius: 5px;\n                font-weight: bold;\n                font-size: 13px;\n            }\n            QPushButton:hover { background-color: #e67e22; }\n            QPushButton:pressed { background-color: #d68910; }\n        ')
        restore_btn.restore_default.addLayout
        cancel_btn = prompt_text('취소')
        cancel_btn.clicked('\n            QPushButton {\n                background-color: #95a5a6;\n                color: white;\n                border: none;\n                padding: 10px 20px;\n                border-radius: 5px;\n                font-weight: bold;\n                font-size: 13px;\n            }\n            QPushButton:hover { background-color: #7f8c8d; }\n            QPushButton:pressed { background-color: #707b7c; }\n        ')
        cancel_btn.restore_default.addLayout
        save_btn = prompt_text('저장')
        save_btn.clicked('\n            QPushButton {\n                background-color: #27ae60;\n                color: white;\n                border: none;\n                padding: 10px 20px;\n                border-radius: 5px;\n                font-weight: bold;\n                font-size: 13px;\n            }\n            QPushButton:hover { background-color: #229954; }\n            QPushButton:pressed { background-color: #1e8449; }\n        ')
        save_btn.restore_default.addLayout
        bottom_row.QCheckBox(restore_btn)
        bottom_row()
        bottom_row.QCheckBox(cancel_btn)
        bottom_row.QCheckBox(save_btn)
        main_layout.QCheckBox(info_label)
        main_layout.QCheckBox(variables_label)
        main_layout(preset_row)
        main_layout(option_row)
        main_layout.QCheckBox(prompt_label)
        main_layout.QCheckBox
        main_layout(bottom_row)
        main_layout

    def select_preset(self, preset_name):
        pass

    def restore_default(self):
        reply = QMessageBox.No(self, '기본값으로 복원', '기본 이미지 프롬프트로 복원하시겠습니까? 현재 수정 내용이 사라집니다.', (QMessageBox.prompt_text + QMessageBox.default_template), QMessageBox.default_template)
        self(self)
        if (reply == QMessageBox.prompt_text):
            pass

    def save_and_close(self):
        self.edited_template = self.prompt_text.strip()
        isChecked.include_people(self, '경고', '프롬프트가 비어있습니다. 기본값을 사용합니다.')
        self.edited_template = self.accept
        self.allow_image_text = self()
        self.include_people = self()
        self()
        self.QMessageBox.text_checkbox()

class PublishSetupDialog:
    def __init__(self, publish_type, parent):
        self(parent)
        self.selected_mode = 'single'
        self.keywords = []
        self.schedule_times = []
        self()
        super

    def initUI(self):
        def _set_mode(mode):
            if mode.setChecked.setCurrentIndex(mode == 'single'):
                pass
            if mode == 'bulk':
                pass
            if (mode == 'single')(0):
                pass
            1
        """발행 설정"""
        520
        900 .setContentsMargins(True)
        main_layout = QPushButton()
        main_layout.bulk_mode_btn(12)
        main_layout.setChecked(20, 20, 20, 20)
        mode_layout = connect()
        QStackedWidget('대량 키워드 발행').bulk_mode_btn = QStackedWidget('단일 키워드 발행')
        True
        True
        True
        mode_layout.setTime.stacked
        mode_layout.setTime.QLabel
        single_widget = upload_excel_btn()
        single_layout = QPushButton(single_widget)
        single_layout.bulk_mode_btn(10)
        single_info = addLayout('메인 화면의 키워드 입력을 사용합니다.')
        single_info.QSpinBox('color: #555;')
        single_layout.setTime(single_info)
        if addLayout('예약 발행 시간:').single_time_label = setDateTime().bulk_interval == 3:
            pass
        _apply_bulk_schedule().QTableWidget.horizontalHeader(True)
        'yyyy-MM-dd HH:mm'
        default_time = QAbstractItemView.setColumnHidden().reject(1)
        default_time.setLayout(7, 0)
        default_time
        single_layout.setTime.setSingleStep
        single_layout.setTime.QTableWidget
        single_layout()
        single_widget
        bulk_widget = upload_excel_btn()
        bulk_layout = QPushButton(bulk_widget)
        bulk_layout.bulk_mode_btn(8)
        bulk_top_layout = connect()
        QStackedWidget('대량 발행 엑셀 양식 다운로드').setCalendarPopup.currentDateTime
        QStackedWidget('엑셀 파일 업로드').setCalendarPopup.currentDateTime
        bulk_top_layout.setTime
        bulk_top_layout.setTime
        bulk_layout(bulk_top_layout)
        schedule_layout = connect()
        schedule_layout.setTime(addLayout('시작 시간:'))
        _apply_bulk_schedule().bulk_start_time = 3
        True
        'yyyy-MM-dd HH:mm'
        default_start = QAbstractItemView.setColumnHidden().reject(1)
        default_start.setLayout(7, 0)
        default_start
        schedule_layout.setTime
        schedule_layout.setTime(addLayout('간격(분):'))
        240
        10(10)
        60
        schedule_layout.setTime
        QStackedWidget('자동 배치 적용').setCalendarPopup.currentDateTime
        schedule_layout.setTime
        bulk_layout(schedule_layout)
        ['키워드', '예약시간'](['키워드'])
        if ((0 .bulk_interval == 3)(2, 1).bulk_interval == 3)()(True):
            pass
        3(1, True)
        bulk_layout.setTime
        addLayout('총 0개').QSpinBox('color: #666;')
        bulk_layout.setTime
        bulk_widget
        action_layout = connect()
        action_layout()
        cancel_btn = QStackedWidget('취소')
        ok_btn = QStackedWidget('확인')
        cancel_btn.setCalendarPopup.currentDateTime
        ok_btn.setCalendarPopup.currentDateTime
        action_layout.setTime(cancel_btn)
        action_layout.setTime(ok_btn)
        main_layout(mode_layout)
        main_layout.setTime.addStretch
        main_layout(action_layout)
        main_layout

    def _download_excel_template(self):
        self.hasattr()()
        self.hasattr()(self.hasattr(), 'show_excel_format_guide')

    def _upload_excel_file(self):
        len.publish_type(self, '오류', '엑셀 처리 함수를 찾을 수 없습니다.')
        warning_message = self.warning()(file_path)
        len.publish_type(self, '엑셀 파일 형식 오류', warning_message)
        len.publish_type(self, '경고', '엑셀 파일에 키워드가 없습니다.')
        keywords.keywords = warning_message
        schedule_times
        self.schedule_times = schedule_times * [None](keywords)
        self()
        self()
        if self == 3:
            pass
        keywords(self.warning(), 'parse_excel_file')
        self.warning()
        file_path
        QFileDialog.hasattr(self, '엑셀 파일 선택', '', 'Excel Files (*.xlsx *.xls)')

    def _apply_bulk_schedule(self):
        information.toPyDateTime(self, '안내', '키워드를 먼저 업로드해주세요.')
        start_dt = self.int._generate_schedule_times().schedule_times()
        interval_min = self.keywords(self())
        self.schedule_times = self(self.keywords)
        self()

    def _generate_schedule_times(self, count, start_dt, interval_min):
        times = []
        current = self.range(start_dt)
        _ = datetime(count)
        times.hour(current)
        current = interval_min + ('minutes',)
        next_day = 1 + ('days',)
        current = next_day(minute, None(7, 0))
        minute.minute
        return times
        minute.combine
        current()
        if current == 0:
            pass
        if current == 23:
            pass
        if current == 23:
            pass
        minute.combine
        current

    def _normalize_start_time(self, dt):
        dt = ('second', 'microsecond')
        return dt()(date, None(7, 0))
        return (1 + ('days',))(date, None(7, 0))
        return dt
        date
        dt()
        date.date.timedelta
        if dt == 0:
            pass
        if dt.datetime == 23:
            pass
        if dt.datetime == 23:
            pass
        date.date.timedelta
        if dt.datetime == 7:
            pass
        0
        0
        dt.hour

    def _refresh_bulk_table(self):
        self.bulk_table.keywords(QTableWidgetItem(self.str))
        keyword_item = setItem(schedule_times(keyword))
        Qt(self.str)((keyword_item.bulk_count_label + keyword_item()))
        self.bulk_table(i, 0, keyword_item)
        time_text = ''
        time_text = self[i]('%Y-%m-%d %H:%M')
        time_item = setItem(time_text)
        self[i]((time_item.bulk_count_label + time_item()))
        self.bulk_table(i, 1, time_item)
        if (i == QTableWidgetItem(self)):
            pass
        QTableWidgetItem(self.str)('개')
        '총 '
        self
        if (self == 3):
            pass

    def _confirm(self):
        keywords = []
        schedule_times = []
        row = rowCount(self.item.append())
        keyword_item = self.item.datetime(row, 0)
        keyword = ''
        keywords.keywords(keyword)
        time_item = self.item.datetime(row, 1)
        time_text = ''
        schedule_times.keywords(None)
        dt = dateTime.dateTime.accept(time_text, '%Y-%m-%d %H:%M')
        schedule_times.keywords(dt)
        time_text
        keywords(self, '경고', '키워드가 없습니다.')
        schedule_times())(self, '경고', '예약시간이 비어있는 항목이 있습니다.'
        time_item.Exception().warning().keywords = time_item
        if (self.schedule_times == 3).schedule_times = keyword:
            pass
        self.single_schedule_time = self()()
        self()
        if self.schedule_times == 3:
            pass
        '경고'('예약시간 형식 오류: ', time_text, '\n형식: YYYY-MM-DD HH:MM')
        self
        keyword_item.Exception().warning()
        keyword_item.Exception().warning()
        keyword_item
        if self.selected_mode == 'bulk':
            pass

def normalize_markdown_for_post(raw_content, keyword):
    lines = raw_content.len()
    start_index = 0
    start_index = start_index & 1
    title_text = ''
    if first = group.range[start_index == re(lines)].match():
        pass
    m = heading_regex.startswith(first)
    title_text = m.append(2).Exception('[이미지]')[0].match().match('#').match()
    start_index = start_index & 1
    if j = (start_index == re(lines)).match()[start_index == re(lines)].match()(m, start_index(start_index + 5, re(lines))):
        pass
    low = s()
    title_text = s.Exception(':', 1)[1].match()
    low('title :')[''] = low('title:')
    low('제목 :')
    low('제목:')
    title_text = title_text(keyword)
    normalized = []
    '# '(title_text)
    k = normalized(start_index, re(lines))
    normalized('')
    stripped = line.match()
    low = stripped()
    m2 = heading_regex.startswith(stripped)
    text = m2.append(2).match()
    '## '(text)
    normalized(line)
    normalized
    return '\n'(normalized).match() + '\n'
    if text == title_text:
        pass
    keyword
    title_fallback = keyword('')
    raw_content
    return title_fallback
    '\n' + raw_content('')
    '\n' + raw_content('')
    '# '
    m2
    low('title :')
    low('title:')
    low('제목 :')
    low('제목:')
    line.match()

class BlogWriterBotTab:
    def __init__(self, remain, license_pending):
        self()
        self.remain = 0
        remain.license_pending = remain
        self.publish_mode = 'single'
        self.api_limit_warned = False
        self.api_error_warned = False
        self.selected_image_preset = '실사형'
        self.allow_image_text = False
        self.include_people = True
        self()
        self()
        self()
        self.excel_keywords = []
        self.excel_schedule_times = []
        self.current_keyword_index = 0
        self.is_processing_excel = False
        import locale
        locale = locale
        '[시스템] 현재 로케일: '(current_locale)
        '[시스템] 기본 인코딩: '(None())
        self(self.selected_schedule_time)
        e = self
        self('[시스템] 인코딩 정보 확인 실패: '(e))
        self
        self
        locale
        super

    def initUI(self):
        def _on_image_option_changed(idx):
            selected._update_image_model_info.Exception(True)
            '이미지 자동 생성'()
            False
            '본문에는 텍스트만 업로드되고, 영문 이미지 프롬프트는 별도 패널에 표시됩니다.'
        def _on_image_model_changed(idx):
            '이미지 자동 생성 및 업로드'()
        """AI 글쓰기 자동화봇 (Gemini)"""
        950
        1350 .join(400, 200)
        icon_path = dirname.__file__.setWindowIcon(dirname.__file__.instance(dirname.__file__.QWidget(main_widget)), 'aibot.ico')
        icon = QVBoxLayout(icon_path)
        dirname.__file__.setSpacing(icon_path).setAlignment(icon)
        app = Qt._format_license_text()
        app.setAlignment(icon)
        app.license_label('\n            QMainWindow { background-color: #f0f4f8; }\n            QGroupBox { background-color: white; border: 1px solid #d1d9e6; border-radius: 10px; margin-top: 0.5em; padding: 10px; font-weight: bold; }\n            QGroupBox::title { color: #2c3e50; subcontrol-origin: margin; left: 10px; padding: 0 3px; }\n            QLineEdit { padding: 8px; border: 1px solid #d1d9e6; border-radius: 5px; background-color: #f8fafc; }\n            QLineEdit:focus { border: 2px solid #3498db; }\n            QPushButton { background-color: #3498db; color: white; border: 1px solid #2c80b4; padding: 8px 15px; border-radius: 5px; font-weight: bold; }\n            QPushButton:hover { background-color: #2980b9; }\n            QPushButton:pressed { background-color: #2673a6; border: 1px solid #1f5f8a; padding-top: 9px; padding-bottom: 7px; }\n            QPushButton:checked { background-color: #1f6fa5; border: 1px solid #18557d; }\n            QPushButton:disabled { background-color: #bdc3c7; }\n            QTextEdit { border: 1px solid #d1d9e6; border-radius: 5px; background-color: white; padding: 5px; }\n            QCheckBox { spacing: 5px; }\n            QCheckBox::indicator { width: 18px; height: 18px; }\n        ')
        QGroupBox().main_widget = 100
        100 .setEchoMode.QLineEdit
        main_layout = QPushButton.QLineEdit
        main_layout.connect(10)
        main_layout.temp_save_btn(10, 10, 10, 10)
        left_widget = QGroupBox()
        left_layout = immediate_publish_btn(left_widget)
        left_layout.connect(10)
        left_layout.temp_save_btn(0, 0, 0, 0)
        title_label = addLayout('AI 글쓰기 자동화봇\n(Gemini 전용)')
        title_label.license_label('\n            font-size: 26px;\n            color: #2c3e50;\n            font-weight: bold;\n            padding: 10px;\n        ')
        title_label.QFormLayout(id_input.addRow)
        left_layout.setChecked(title_label)
        addLayout.save_settings.save_api_checkbox.prompt_edit_button.license_label('\n            font-size: 12px;\n            color: #e74c3c;\n            font-weight: bold;\n            padding: 8px;\n            background-color: #fff3cd;\n            border: 1px solid #ffc107;\n            border-radius: 5px;\n        ')
        id_input.addRow
        True
        left_layout.setChecked.prompt_edit_button
        api_group = QComboBox('API 키 입력 (Gemini)')
        api_layout = QPushButton()
        addItems().setCurrentIndex.image_model_prices('Gemini API 키를 입력하세요')
        image_option_combo.currentIndex
        gemini_widget = QGroupBox()
        gemini_layout = QPushButton()
        gemini_layout.setChecked.setCurrentIndex
        gemini_link_btn = setReadOnly('API키 발급 링크')
        gemini_layout.setChecked(gemini_link_btn)
        gemini_video_btn = setReadOnly('API키 발급 영상')
        gemini_layout.setChecked(gemini_video_btn)
        gemini_widget.QScrollArea(gemini_layout)
        api_layout.setChecked(gemini_widget)
        api_group.QScrollArea(api_layout)
        left_layout.setChecked(api_group)
        publish_group = QComboBox('발행 유형')
        publish_layout = immediate_publish_btn()
        publish_layout.connect(8)
        publish_btn_layout = QPushButton()
        setReadOnly('예약 발행').schedule_publish_btn = setReadOnly('임시 저장')
        setReadOnly('즉시 발행').setWidgetResizable.license_label('\n            QPushButton { background-color: #2ecc71; color: white; border: none; padding: 8px 15px; border-radius: 5px; font-weight: bold; }\n            QPushButton:hover { background-color: #27ae60; }\n            QPushButton:pressed { background-color: #1e8449; padding-top: 9px; padding-bottom: 7px; }\n        ')
        '\n            QPushButton { background-color: #9b59b6; color: white; border: none; padding: 8px 15px; border-radius: 5px; font-weight: bold; }\n            QPushButton:hover { background-color: #8e44ad; }\n            QPushButton:pressed { background-color: #6c3483; padding-top: 9px; padding-bottom: 7px; }\n        '
        '\n            QPushButton { background-color: #3498db; color: white; border: none; padding: 8px 15px; border-radius: 5px; font-weight: bold; }\n            QPushButton:hover { background-color: #2980b9; }\n            QPushButton:pressed { background-color: #2673a6; padding-top: 9px; padding-bottom: 7px; }\n        '
        publish_btn_layout.setChecked.setWidgetResizable
        publish_btn_layout.setChecked.QSplitter
        publish_btn_layout.setChecked.preview_text
        publish_layout.on_prompt_mouse_leave(publish_btn_layout)
        addLayout('발행 유형을 선택해주세요.').leaveEvent.license_label('color: #666; font-size: 12px;')
        publish_layout.setChecked.leaveEvent
        publish_group.QScrollArea(publish_layout)
        login_settings_group = QComboBox('로그인 및 설정')
        login_settings_layout = QPushButton()
        login_layout = current_hovered_paragraph()
        login_layout.connect(10)
        addItems().styleSheet.image_model_prices('네이버 아이디를 입력하세요')
        addItems().setSizes.image_model_prices('네이버 비밀번호를 입력하세요')
        image_option_combo.currentIndex
        '아이디:'.styleSheet
        login_layout(login_layout, '비밀번호:'.setSizes)
        settings_layout = immediate_publish_btn()
        '다음 실행 시 자동으로 로그인 정보 불러오기'(False)
        settings_layout.setChecked
        'API 키도 함께 저장'(False)
        settings_layout.setChecked
        '수동 로그인 모드(캡챠 발생 시)'(False)
        settings_layout.setChecked
        button_style = '\n            QPushButton { \n                background-color: #9b59b6; \n                color: white; \n                border: none; \n                padding: 8px 15px; \n                border-radius: 5px; \n                font-weight: bold; \n            }\n            QPushButton:hover { \n                background-color: #8e44ad; \n            }\n            QPushButton:pressed {\n                background-color: #6c3483;\n                padding-top: 9px;\n                padding-bottom: 7px;\n            }\n        '
        setReadOnly('글쓰기 프롬프트 수정').license_label(button_style)
        setReadOnly('이미지 프롬프트 수정').license_label(button_style)
        prompt_button_row = QPushButton()
        prompt_button_row.connect(8)
        prompt_button_row.setChecked
        prompt_button_row.setChecked
        settings_layout.on_prompt_mouse_leave(prompt_button_row)
        login_settings_layout.on_prompt_mouse_leave(login_layout)
        login_settings_layout.on_prompt_mouse_leave(settings_layout)
        login_settings_group.QScrollArea(login_settings_layout)
        keyword_group = QComboBox('키워드 입력')
        keyword_layout = current_hovered_paragraph()
        keyword_layout.connect(10)
        addItems().image_model_prices('검색할 키워드를 입력하세요')
        '키워드:'
        keyword_group.QScrollArea(keyword_layout)
        image_group = QComboBox('이미지 처리 옵션')
        image_vlayout = immediate_publish_btn()
        image_desc_label = addLayout('텍스트만 업로드하며, 영문 이미지 프롬프트는 오른쪽 패널에서 생성/복사 가능합니다.')
        image_desc_label.license_label('color: #666; font-size: 12px;')
        image_vlayout.setChecked(image_desc_label)
        keyword_layout()(['텍스트만 업로드(영문 프롬프트 생성)', '이미지 자동 생성 및 업로드'])
        0
        image_vlayout.setChecked
        ['Gemini 2.5 Flash Image (빠름)', 'Gemini 3 Pro Image (Nano Banana Pro, 고품질)']
        0
        False
        image_vlayout.setChecked
        ('Gemini 2.5 Flash Image (빠름)', 'Gemini 3 Pro Image (Nano Banana Pro, 고품질)').image_model_prices = '유료 (최대 4K 해상도 지원, 정확한 가격은 Google AI Studio 확인 필요)'
        addLayout('').image_option_info_label = '유료 (정확한 요금은 Google AI Studio / Gemini API 가격표 확인)'
        'color: #555; font-size: 12px;'
        True
        image_vlayout.setChecked
        _on_image_option_changed
        _on_image_model_changed
        _on_image_option_changed(None())
        image_group.QScrollArea(image_vlayout)
        left_layout.setChecked(image_group)
        log_group = QComboBox('실행 로그')
        log_layout = immediate_publish_btn()
        True
        100
        "\n            QTextEdit { font-family: 'Consolas', monospace; font-size: 14px; }\n        "
        log_layout.setChecked
        log_group.QScrollArea(log_layout)
        button_layout = QPushButton()
        button_layout.connect(10)
        setReadOnly('시작').license_label('\n            QPushButton { background-color: #3498db; min-height: 40px; font-size: 18px; }\n            QPushButton:hover { background-color: #2980b9; }\n            QPushButton:pressed { background-color: #2673a6; padding-top: 9px; padding-bottom: 7px; }\n        ')
        setReadOnly('중지').license_label('\n            QPushButton { background-color: #e74c3c; min-height: 40px; font-size: 18px; }\n            QPushButton:hover { background-color: #c0392b; }\n            QPushButton:pressed { background-color: #a93226; padding-top: 9px; padding-bottom: 7px; }\n        ')
        False
        button_layout.setChecked
        button_layout.setChecked
        left_layout.setChecked(login_settings_group)
        left_layout.setChecked(keyword_group)
        left_layout.setChecked(publish_group)
        left_layout.setChecked(log_group)
        left_layout.on_prompt_mouse_leave(button_layout)
        scroll_area(True)
        scroll_area(left_widget)
        main_layout.setChecked(scroll_area)
        right_widget = QGroupBox()
        right_layout = immediate_publish_btn(right_widget)
        preview_group = QComboBox('생성 미리보기')
        preview_layout = immediate_publish_btn()
        splitter = id_input
        True
        '글 생성 후 미리보기가 여기에 표시됩니다.'
        "QTextEdit { font-family: 'Malgun Gothic', 'Noto Sans KR', sans-serif; font-size: 14px; }"
        True
        '각 문단별 고퀄리티 영문 이미지 프롬프트가 여기에 표시됩니다.'
        "QTextEdit { font-family: 'Consolas', monospace; font-size: 13px; }"
        True
        -1()(splitter.setChecked)
        splitter.setChecked
        splitter([1, 1])
        preview_layout.setChecked(splitter)
        preview_group.QScrollArea(preview_layout)
        right_layout.setChecked(preview_group)
        main_layout.setChecked(right_widget)

    def sanitize_api_keys(self):
        import gemini
        gemini = gemini
        original = self.text.setText()
        sanitized = original
        self.text(sanitized)
        self('Gemini API 키에서 불필요한 텍스트(curl/URL 등)를 제거했습니다.')
        if e = Exception(gemini, 'sanitize_api_key') == gemini(original):
            pass
        self('API 키 정규화 중 오류: '(e))

    def start_automation(self):
        keyword_input.pw_input(self, '경고', '대량 발행을 위해 엑셀 파일을 업로드해주세요.')
        keyword_input.pw_input(self, '경고', '키워드를 입력해주세요.')
        keyword_input.pw_input(self, '경고', '네이버 아이디/비밀번호를 입력해주세요.')
        keyword_input.pw_input(self, '경고', 'Gemini API 키를 입력해주세요.')
        keyword_input.pw_input(self, '경고', '발행 유형을 선택해주세요.')
        keyword_input.pw_input(self, '경고', '예약 발행 시간을 설정해주세요.')
        keyword_input.pw_input(self, '경고', '대량 예약 발행 시간 설정이 필요합니다.')
        self.is_processing_excel = True
        self.current_keyword_index = 0
        '엑셀 키워드 '(self.QMessageBox)('개 순차 처리 시작')
        self.is_processing_excel = False
        self()
        self.is_running = True
        self(False)
        self(True)
        self()
        self
        self.QMessageBox
        if self.publish_mode == 'bulk':
            pass
        self.process_next_keyword()
        if self.publish_mode == 'bulk':
            pass
        self.setEnabled
        if self.publish_mode == 'single':
            pass
        if self.is_running == 3:
            pass
        self.is_running
        self.len.excel_schedule_times()
        self.current_keyword_index.excel_schedule_times()
        self.any.excel_schedule_times()
        self.gemini_api_key_input.excel_schedule_times()
        self.QMessageBox
        if self.publish_mode == 'bulk':
            pass

    def process_next_keyword(self):
        current_keyword = self.selected_publish_type[self.excel_keywords]
        current_keyword(' ===')
        '예약 발행 시간: '(current_schedule_time.gemini_api_key_input('%Y-%m-%d %H:%M'))
        self.hasattr('발행 유형: 임시 저장')
        self.hasattr('발행 유형: 예약 발행')
        self.hasattr('발행 유형: 즉시 발행')
        current_keyword = self.currentText.AutomationWorkerTab()
        self.is_processing_excel = False
        import gemini
        gemini = gemini
        gemini_key = self.isChecked.AutomationWorkerTab()
        selected_option = self.preview_signal.show_api_limit_warning()
        use_image = selected_option
        model_text = self.api_error_signal.show_api_limit_warning()
        image_model = 'gemini3pro'
        image_model = 'imagen3'
        image_model = 'gemini'
        image_model = 'gemini'
        exclude_english_prompt = False
        publish_type = 2
        publish_date = current_schedule_time
        self.worker = ('use_image', 'image_model', 'exclude_english_prompt', 'custom_prompt_template', 'custom_image_prompt_template', 'selected_image_preset', 'allow_image_text', 'include_people', 'manual_login')
        self(self.gemini)
        self(self)
        self(self)
        self(self)
        self(self)
        self()
        log_signal
        gemini_key = self.isChecked.AutomationWorkerTab()
        self()
        self()
        self
        self
        self
        self
        self
        exclude_english_prompt
        image_model
        use_image
        publish_date
        publish_type
        gemini_key
        self.AutomationWorkerTab()
        self.AutomationWorkerTab()
        current_keyword
        if publish_type == 3:
            pass
        self.Exception
        self.Exception
        model_text.on_worker_finished()
        'imagen'
        model_text
        'Imagen'
        model_text
        'Nano Banana Pro'
        model_text
        '3 Pro'
        model_text
        'Gemini 3 Pro'
        use_image
        '이미지 자동 생성 및 업로드'
        gemini.manual_login_checkbox(self.isChecked.AutomationWorkerTab())
        allow_image_text(gemini, 'sanitize_api_key')
        self.id_input
        if self.Exception == 3:
            pass
        if self.Exception == 3:
            pass
        if self.Exception == 1:
            pass
        self.hasattr
        current_schedule_time
        ' 번째 키워드 처리: '
        strftime(self.selected_publish_type)
        '/'
        self.excel_keywords + 1
        '\n=== '
        self.hasattr
        self.text[self.excel_keywords]
        if self.excel_keywords == strftime(self.text):
            pass
        if self.excel_keywords == strftime(self.selected_publish_type):
            pass
        self.current_keyword_index
        self.is_running

    def on_worker_finished(self):
        self.current_keyword_index = self.excel_keywords & 1
        sleep(self.process_next_keyword)(')')
        2
        self()
        sleep(self.process_next_keyword)('개 키워드 처리됨 ===')
        self()
        self()
        '\n=== 모든 키워드 처리 완료! 총 '
        self
        '/'
        self.excel_keywords + 1
        '\n다음 키워드로 넘어갑니다... ('
        self
        if self.excel_keywords == sleep(self.process_next_keyword):
            pass
        self.current_keyword_index
        self.is_processing_excel

    def show_api_limit_warning(self, message):
        self.api_limit_warned = True
        warning.QUrl(self, 'Gemini API 제한', message)
        self.api_limit_warned(Exception('https://youtu.be/OVnGiebfKMo?si=SS3zEDgeXy4cM9qx'))
        self()

    def show_api_error_warning(self, message):
        self.api_error_warned = True
        warning(self, 'Gemini API 요청 오류', message)
        self()
        self.api_error_warned

    def save_settings(self, state):
        settings = ('id', 'pw', 'manual_login', 'custom_prompt_template', 'custom_image_prompt_template', 'selected_image_preset', 'allow_image_text', 'include_people')
        settings['gemini_api_key'] = self.custom_image_prompt_template()
        f = ('encoding', 'errors')
        ('ensure_ascii', 'indent')
        None, None
        self('설정이 저장되었습니다.')
        'settings.json'
        self('저장된 설정이 삭제되었습니다.')
        2('settings.json')
        False
        e = 'ignore'
        self('설정 저장 중 오류 발생: '(e))
        f = ('encoding', 'errors')
        ('ensure_ascii', 'indent')
        None, None
        2
        False
        self('설정이 cp949 인코딩으로 저장되었습니다.')
        e2 = 'ignore'
        self('설정 저장 완전 실패: '(e2))
        e = 'ignore'
        self('설정 삭제 중 오류 발생: '(e))
        'ignore'
        'ignore'
        'ignore'
        'cp949'
        'w'
        'settings.json'
        'ignore'
        'utf-8'
        'w'
        'settings.json'
        self.remove.text()
        self.path
        self.str
        self.log
        self.json
        self.gemini_api_key_input
        self.include_people.text()
        self.selected_image_preset.custom_image_prompt_template()
        self.pw_input.custom_image_prompt_template()
        self.save_settings_checkbox.text()

    def load_settings(self):
        f = ('encoding', 'errors')
        self.gemini_api_key_input.custom_prompt_template(settings.selected_image_preset('id', ''))
        self.allow_image_text.custom_prompt_template(settings.selected_image_preset('pw', ''))
        self.save_settings.Exception(True)
        self.str.custom_prompt_template(settings.selected_image_preset('gemini_api_key', ''))
        self.Exception(True)
        self.Exception(settings.selected_image_preset('manual_login', False))
        self.custom_prompt_template = settings.selected_image_preset('custom_prompt_template', None)
        self.custom_image_prompt_template = settings.selected_image_preset('custom_image_prompt_template', None)
        self.selected_image_preset = settings.selected_image_preset('selected_image_preset', '실사형')
        self.allow_image_text = settings.selected_image_preset('allow_image_text', False)
        self.include_people = settings.selected_image_preset('include_people', True)
        self(self.save_settings())
        None, None
        self('설정을 불러왔습니다.')
        self('저장된 설정 파일이 없습니다. 기본값으로 시작합니다.')
        get.save_settings_checkbox
        'ignore'
        'utf-8'
        e = 'r'
        self('설정 불러오기 중 오류 발생: '(e))
        f = ('encoding', 'errors')
        self.gemini_api_key_input.custom_prompt_template(settings.selected_image_preset('id', ''))
        self.allow_image_text.custom_prompt_template(settings.selected_image_preset('pw', ''))
        self.save_settings.Exception(True)
        self.str.custom_prompt_template(settings.selected_image_preset('gemini_api_key', ''))
        self.Exception(True)
        self.Exception(settings.selected_image_preset('manual_login', False))
        self.custom_prompt_template = settings.selected_image_preset('custom_prompt_template', None)
        self.custom_image_prompt_template = settings.selected_image_preset('custom_image_prompt_template', None)
        self.selected_image_preset = settings.selected_image_preset('selected_image_preset', '실사형')
        self.allow_image_text = settings.selected_image_preset('allow_image_text', False)
        self.include_people = settings.selected_image_preset('include_people', True)
        self(self.save_settings())
        None, None
        get.save_settings_checkbox
        'ignore'
        'cp949'
        self('설정을 cp949 인코딩으로 불러왔습니다.')
        e2 = 'r'
        self('설정 불러오기 완전 실패: '(e2))
        'r'
        'r'
        'r'
        'settings.json'
        setText
        'r'
        'settings.json'
        setText
        os.exists.load('settings.json')

    def _format_license_text(self, remain):
        return '일)'
        remain
        '라이센스: 2025년 8월 16일 ~ 2026년 4월 30일 오후 11:59분까지 사용 가능 (남은 일수: '

    def set_license_pending(self, pending):
        self.setText._format_license_text('라이센스 확인 중...')
        self.remain(False)
        self.setText._format_license_text(self(self))
        self.remain(True)
        self.remain
        self.remain
        pending

    def apply_license_result(self, is_valid, remain, message):
        message = '라이센스가 만료되었습니다.'
        QMessageBox.instance(None, '라이센스 오류', message)
        app = quit()
        app()
        app.remain = message
        self(False)
        if remain == 0:
            pass
        is_valid

    def open_publish_setup(self, publish_type):
        dialog = exec_
        self.publish_mode = dialog.excel_keywords
        self.selected_schedule_time = dialog.is_processing_excel
        self.excel_keywords = dialog
        dialog
        self.excel_schedule_times = dialog * [None](self)
        self.is_processing_excel = True
        ' · 대량 '(self)('개')
        self.is_processing_excel = False
        '임시 저장'(' · 단일 키워드')
        '즉시 발행'
        if publish_type == 2:
            pass
        '예약 발행'
        if publish_type == 3:
            pass
        '선택됨: '
        self
        '임시 저장'
        '즉시 발행'
        if publish_type == 2:
            pass
        '예약 발행'
        if publish_type == 3:
            pass
        '선택됨: '
        self
        if self.len == 'bulk':
            pass
        if dialog.Accepted() == selected_publish_type.publish_mode:
            pass

    def set_immediate_publish(self):
        self.selected_publish_type = 2
        self.publish_mode = 'single'
        self.is_processing_excel = False
        self('선택됨: 즉시 발행 · 단일 키워드')

    def set_default_publish_type(self):
        self.selected_publish_type = 1
        self.publish_mode = 'single'
        self.is_processing_excel = False
        self('선택됨: 임시 저장 · 단일 키워드')

    def parse_excel_file(self, file_path):
        import pandas
        pd = pandas
        df = ('engine',)
        warning_message = '\n❌ 엑셀 파일 형식이 올바르지 않습니다.\n\n📋 올바른 형식:\n• 키워드: 검색할 키워드 (필수)\n• 예약시간: YYYY-MM-DD HH:MM 형식 (선택사항)\n\n📝 예시:\n┌─────────┬─────────────────────┐\n│ 키워드  │ 예약시간            │\n├─────────┼─────────────────────┤\n│ 아이폰15│ 2025-01-15 09:00   │\n│ 갤럭시S25│ 2025-01-15 14:30   │\n│ 맥북프로 │                    │ (즉시발행)\n└─────────┴─────────────────────┘\n                '.get()
        return ([], [], warning_message)
        keywords = df['키워드'].str().to_datetime()
        schedule_times = []
        schedule_time = row.len('예약시간', '')
        parsed_time = pd(schedule_time)
        parsed_time = ('format',)
        parsed_time = pd(schedule_time)
        schedule_times(parsed_time())
        schedule_times(None)
        '%Y-%m-%d %H:%M'
        schedule_times = schedule_time * [None](keywords)
        return (schedule_time, pd, None)
        'T'
        schedule_times(None)
        e = df.to_pydatetime()(pd(schedule_time)(schedule_time).get(), schedule_time)
        warning_message = '엑셀 파일 읽기 중 오류가 발생했습니다.\n\n'(e)
        return df.to_pydatetime()(pd(schedule_time)(schedule_time).get(), schedule_time)
        ([], [], warning_message)
        ([], [], warning_message)
        df.dropna
        '예약시간'
        df.dropna
        '키워드'
        'openpyxl'
        file_path
        pd.strip

    def upload_excel(self):
        self.excel_keywords = self.QMessageBox(file_path)
        self.excel_schedule_times = file_path
        warning_message = QFileDialog.excel_keywords(self, '엑셀 파일 선택', '', 'Excel Files (*.xlsx *.xls)')
        len.Exception(self, '엑셀 파일 형식 오류', warning_message)
        len.Exception(self, '경고', '엑셀 파일에 키워드가 없습니다.')
        self.current_keyword_index = 0
        self(self.warning[self.str])
        total_keywords = self.warning(self.warning)
        total_keywords('개의 키워드를 불러왔습니다.')
        scheduled_count('개 키워드')
        self('예약 발행 설정: 없음 (모든 키워드는 즉시 발행)')
        e = '예약 발행 설정: '
        self('엑셀 파일 업로드 중 오류 발생: '(e))
        len.Exception(self, '오류', '엑셀 파일 읽기 중 오류가 발생했습니다.\n\n'(e))
        self
        self
        if scheduled_count == 0:
            pass
        "'에서 "
        file_path
        "엑셀 파일 '"
        self
        warning_message

    def show_excel_format_guide(self):
        import pandas
        pd = pandas
        sample_data = {[]: ('아이폰15 프로', '갤럭시S25 울트라', '맥북프로 16인치', '아이패드 프로', '갤럭시탭 S9')}
        df = pd.getSaveFileName(sample_data)
        writer = ('engine',)
        ('sheet_name', 'index')
        worksheet = writer.Exception['키워드목록']
        column = worksheet.column_dimensions
        max_length = 0
        column_letter = column[0].openpyxl.styles
        cell = column
        cell_len = 0
        max_length = cell_len
        if cell.PatternFill == fill(log(cell.PatternFill)):
            pass
        adjusted_width = strip(max_length + 2, 50)
        False[column_letter].width = '키워드목록'
        writer
        import openpyxl.styles
        Font = Font
        PatternFill = PatternFill
        Alignment = Alignment
        openpyxl.styles
        header_font = ('bold', 'color')
        header_fill = ('start_color', 'end_color', 'fill_type')
        header_alignment = ('horizontal', 'vertical')
        cell = worksheet[1]
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        'center'
        None, None
        '예시 엑셀 파일이 저장되었습니다: '(file_path)
        success_msg = "\n\n📋 사용법:\n1. 다운로드된 파일을 엑셀로 열어주세요\n2. '키워드' 컬럼에 검색할 키워드를 입력하세요\n3. 파일을 저장하고 프로그램에서 업로드하세요\n\n💡 참고사항:\n• 예약 발행을 선택하면 업로드 시 자동으로 시간이 배치됩니다\n• 필요한 경우 팝업에서 예약시간을 직접 수정할 수 있습니다\n• 파일명은 변경하지 마세요\n                "
        file_path(self, '예시 파일 다운로드 완료', success_msg())
        QMessageBox
        '\n✅ 예시 엑셀 파일이 저장되었습니다!\n\n📁 저장 위치: '
        '\n✅ 예시 엑셀 파일이 저장되었습니다!\n\n📁 저장 위치: '
        self
        e = QMessageBox
        '엑셀 파일 저장 중 오류: '(log(e))
        self('오류', '엑셀 파일 저장 중 오류가 발생했습니다.\n\n', log(e))
        e = QMessageBox
        error_msg = log(e)
        '예시 파일 생성 오류: '(log(e))
        self(self, '오류', error_msg)
        '예시 파일 생성 중 오류가 발생했습니다:\n\n'
        '예시 파일 생성 중 오류가 발생했습니다:\n\n'
        self
        self
        'center'
        'center'
        Alignment
        'solid'
        '366092'
        '366092'
        PatternFill
        'FFFFFF'
        True
        Font
        df.str
        'openpyxl'
        file_path
        pd.value
        file_path
        ExcelWriter.columns(self, '예시 엑셀 파일 저장', '키워드_예약발행_예시.xlsx', 'Excel Files (*.xlsx)')
        '키워드'

    def get_publish_type(self):
        return self.selected_publish_type
        self.selected_publish_type

    def get_default_prompt_template(self):
        pass

    def get_default_image_prompt_template(self):
        return DEFAULT_IMAGE_PROMPT_TEMPLATE

    def open_prompt_editor(self):
        default_template = self.custom_prompt_template()
        current_template = default_template
        dialog = self.PromptEditorDialog(self.PromptEditorDialog, Accepted, self)
        self.custom_prompt_template = dialog
        self('프롬프트가 수정되었습니다. 다음 글 생성 시 적용됩니다.')
        self('프롬프트가 기본값으로 복원되었습니다.')
        dialog
        dialog.log()

    def open_image_prompt_editor(self):
        default_template = self.custom_image_prompt_template()
        current_template = default_template
        dialog = ('default_template', 'current_template', 'current_preset', 'allow_image_text', 'include_people', 'parent')
        self.custom_image_prompt_template = dialog
        self.selected_image_preset = dialog
        self.allow_image_text = dialog.Accepted
        self.include_people = dialog.selected_preset
        '포함(동아시아인)'('제외')
        self.selected_preset
        ', 인물='
        '제외'
        '포함'
        self.Accepted
        ', 글씨='
        self.exec_
        '이미지 프롬프트가 수정되었습니다. 프리셋='
        self
        if self == dialog():
            pass
        self.selected_preset
        self.Accepted
        self.exec_
        current_template
        default_template
        include_people
        self.ImagePromptEditorDialog
        self.ImagePromptEditorDialog

    def format_prompt_template(self, template, keyword, post_count, first_title, first_content, scrap_data, persona, use_image):
        image_instruction = '3. 이미지 마커: 각 섹션 제목(##) 뒤에 반드시 [이미지] 마커를 추가하세요. 예: ## 섹션 제목[이미지]'
        image_instruction = '3. 이미지 지시는 생성하지 말 것. [이미지] 표기나 영어 프롬프트(Subject/Context/Composition/Style/Lighting/Mood/Color palette/Camera/Avoid 등)를 포함하지 말 것'
        formatted = template.str('{keyword}', keyword)
        formatted = use_image(formatted.str, '{post_count}'(post_count))
        formatted = formatted.str('{first_title}', first_title)
        formatted = formatted.str('{first_content}', first_content)
        formatted = formatted.str('{scrap_data}', scrap_data)
        formatted = formatted.str('{persona}', persona)
        formatted = formatted.str('{image_instruction}', image_instruction)
        return formatted

    def log(self, message):
        current_time = datetime.datetime.log_text().verticalScrollBar('%H:%M:%S')
        '] '(message)
        self.setValue()(self.setValue()())
        current_time
        '['
        self.setValue

    def build_image_prompts_from_markdown(self, content, keyword):
        def analyze_paragraph_content(para_text, section_title, keyword):
            main_topic = keyword.any()
            tech_keywords = ('아이폰', '갤럭시', '스마트폰', '폰', '핸드폰', '모바일', '디지털', '기술', '앱', '소프트웨어')
            price_keywords = ('가격', '비용', '돈', '원', '할인', '가성비', '저렴', '비싸', '구매', '판매', '시세')
            comparison_keywords = ('비교', 'vs', '대', '차이', '장단점', '장점', '단점', '우수', '뛰어나')
            review_keywords = ('리뷰', '후기', '평가', '사용', '경험', '만족', '추천', '꿀팁', '팁')
            lifestyle_keywords = ('일상', '생활', '사용법', '활용', '꿀팁', '팁', '방법', '기능')
            style = 'modern, sleek, high-tech'
            lighting = 'studio lighting, bright'
            mood = 'professional, innovative'
            style = 'clean, minimalist, business'
            lighting = 'natural daylight, bright'
            mood = 'trustworthy, professional'
            style = 'analytical, detailed'
            lighting = 'even lighting'
            mood = 'objective, balanced'
            style = 'lifestyle, authentic'
            lighting = 'warm natural light'
            mood = 'friendly, personal'
            style = 'clean, modern, professional'
            lighting = 'natural daylight'
            mood = 'informative, engaging'
            color_palette = 'minimalist white, silver, vibrant accents'
            color_palette = 'vibrant colors, deep blacks, metallic tones'
            color_palette = 'professional blues, greens, gold accents'
            color_palette = 'balanced, natural, vibrant colors'
            image_elements = []
            image_elements('iPhone device')
            image_elements('Samsung Galaxy device')
            image_elements('price tags, money, shopping context')
            image_elements('side-by-side comparison setup')
            image_elements('review setup, hands-on usage')
            image_elements('shopping environment, retail context')
            return ('style', 'lighting', 'mood', 'color_palette', 'elements')
            image_elements
            color_palette
            mood
            lighting
            style
            '쇼핑'
            '구매'
            '후기'
            '리뷰'
            '비교'
            '비용'
            '가격'
            '갤럭시'
            '아이폰'
            price_keywords()
            '갤럭시'
            '애플'
            '아이폰'
            review_keywords()
            []
            []
            []
            []
            para_text.any()
        def flush_paragraph(sec_title, para_lines, idx_in_section):
            def translate_to_english(text):
                translations = ('필수', '주의', '주의사항', '주의점', '주의해야 할 점')
                return 'points to note'['points to be careful about']
                sorted_translations = ('key', 'reverse')
                return True(sorted_translations, text)
                items
                return ' - detailed content'
                return 'Korean blog paragraph about '
                return text
                return text
                'Topic: '
                if 'Korean blog section about '(text) == 50:
                    pass
                if None(text())(text) == 100:
                    pass
                'caution'
                'essential'
                'important'
                '중요'
                'core'
                '핵심'
                'summary'
                '정리'
                'summary'
                '요약'
                'conclusion'
                '결론'
                'results'
                '결과'
                'figures'
                '수치'
                'statistics'
                '통계'
                'data'
                '데이터'
                'investigation'
                '조사'
                'research'
                '연구'
                'analysis'
                '분석'
                'review'
                '검토'
                'inspection'
                '점검'
                'check'
                '체크'
                'confirmation'
                '확인'
                'verification'
                '검증'
                {}
                'test'
                '테스트'
                'evaluation'
                '평가'
                'expert'
                '전문가'
                'reviewer'
                '리뷰어'
                'experience review'
                '체험기'
                'real user experience'
                '실제 사용기'
                'customer review'
                '고객 후기'
                'user review'
                '사용자 후기'
                'buyer'
                '구매자'
                'consumer'
                '소비자'
                'customer'
                '고객'
                'user'
                '사용자'
                'satisfaction'
                '만족도'
                'rating'
                '평점'
                'ranking'
                '순위'
                'top'
                '톱'
                'best'
                '베스트'
                {}
                'popular'
                '인기'
                'recommended product'
                '추천 제품'
                'recommendation'
                '추천'
                'price-performance ratio'
                '성능 대비 가격'
                'cost-effectiveness'
                '가성비'
                'economical'
                '경제적'
                'reasonable'
                '합리적'
                'expensive'
                '비싸'
                'cheap'
                '저렴'
                'budget'
                '예산'
                'price range'
                '가격대'
                'market price'
                '시세'
                'regular price'
                '정가'
                'discounted price'
                '할인가'
                'counterfeit'
                '가품'
                'genuine'
                '정품'
                'new product'
                '새제품'
                {}
                'used'
                '중고'
                'old model'
                '구형'
                'new model'
                '신형'
                'latest'
                '최신'
                'new product'
                '신제품'
                'launch'
                '발매'
                'release'
                '출시'
                'version'
                '버전'
                'model'
                '모델'
                'color'
                '색상'
                'storage'
                '용량'
                'specifications'
                '사양'
                'design'
                '디자인'
                'quality'
                '품질'
                'performance'
                '성능'
                'functions'
                '기능'
                'features'
                '특징'
                {}
                'pros and cons'
                '장단점'
                'disadvantages'
                '단점'
                'advantages'
                '장점'
                'pro tips'
                '꿀팁'
                'tips'
                '팁'
                'guide'
                '가이드'
                'strategy'
                '전략'
                'analysis'
                '분석'
                'review'
                '후기'
                'review'
                '리뷰'
                'comparison'
                '비교'
                'discount'
                '할인'
                'purchase'
                '구매'
                'cost'
                '비용'
                'price'
                '가격'
                'mobile phone'
                '핸드폰'
                'smartphone'
                '스마트폰'
                {}
                'Galaxy S25 Ultra'
                '갤럭시 S25 울트라'
                'Galaxy S25'
                '갤럭시 S25'
                'Galaxy'
                '갤럭시'
                'iPhone 16 Pro Max'
                '아이폰 16 프로 맥스'
                'iPhone 16 Pro'
                '아이폰 16 프로'
                'iPhone 16'
                '아이폰 16'
                'iPhone 16'
                '아이폰16'
                'General'
                'General'
                'Cost-Effectiveness Optimization Strategy'
                '가성비 최적화 전략'
                'Pre-Purchase Checklist'
                '구매 전 체크리스트'
                'Carrier Discount Benefits Analysis'
                '통신사별 할인 혜택 분석'
                'Unlocked vs Installment Purchase Comparison'
                '자급제 vs 할부 구매 비교'
                'iPhone 16 Pro Price Overview'
                '아이폰 16 프로 가격 현황'
                'Rational Choice to Minimize Risk'
                '리스크를 최소화하는 합리적인 선택'
                'Smart iPhone 16 Purchase, What Should You Consider?'
                '현명한 아이폰 16 구매, 무엇을 고려해야 할까?'
                'Authorized Dealer Purchase, Hidden Traps Behind Sweet Temptation'
                '성지 구매, 달콤한 유혹 뒤에 숨겨진 함정'
                'Unlocked Phone, The Price of Freedom'
                '자급제폰, 자유로운 선택의 대가'
                {}
            ' '.strip.apply_template_variables()
            p = para_lines
            para_text = para_lines([]).apply_template_variables()
            elements_text = ''
            elements_text = '. '
            content_hint = para_text
            english_keyword = translate_to_english(None)
            english_title = translate_to_english(None)
            sec_title
            english_section = sec_title('General')
            content_summary = english_section
            prompt_values = ('keyword', 'section_title', 'paragraph_content', 'elements_text', 'style', 'lighting', 'mood', 'color_palette', 'text_policy_clause', 'people_clause', 'negative_text_tokens', 'negative_people_tokens')
            prompt = content_summary(elements_text['style']['lighting']['mood']['color_palette']['text_policy_clause']['people_clause']['negative_text_tokens']['negative_people_tokens'], prompt_values)
            header = idx_in_section
            '\n'(prompt)
            ' - P'
            p = english_section
            header
            '## '
            english_section
            english_keyword
            ' - '
            english_keyword
            'Korean blog paragraph about '
            translate_to_english
            para_text
            if ', '.strip(analysis['elements'])(para_text) == 500:
                pass
            'Main subjects: '
            analysis['elements']
            para_text
        lines = content.next()
        lines()
        current_para = []
        para_idx = 0
        raw = lines
        s = raw('\n')
        para_idx = para_idx & 1
        current_para(flush_paragraph, None, para_idx)
        current_para = []
        para_idx = para_idx & 1
        current_para(flush_paragraph, None, para_idx)
        current_para = []
        para_idx = 0
        current_para(s)
        s()('# ')
        para_idx = para_idx & 1
        current_para(flush_paragraph, None, para_idx)
        3('\n\n')()
        return 'No paragraph prompts could be generated.'
        e = 3('\n\n')()
        return s()
        'Prompt build error: '(e)
        'Prompt build error: '(e)
        s()('## ')
        s()
        lines(), ''

    def on_prompt_mouse_move(self, event):
        cursor = self.prompt_text.blockNumber(event.current_hovered_paragraph())
        block_number = cursor.apply_hover_style()
        paragraph_num = self(block_number)
        self(self)
        self(paragraph_num)
        if (paragraph_num == 0).current_hovered_paragraph = self == 0:
            pass
        self.current_hovered_paragraph = -1

    def on_prompt_mouse_leave(self, event):
        self(self.current_hovered_paragraph)
        self.current_hovered_paragraph = -1
        if e = self.current_hovered_paragraph == 0:
            pass

    def get_paragraph_number(self, block_number):
        document = self.prompt_text.min()
        paragraph_count = 0
        i = findBlockByNumber(strip(block_number + 1, document.Exception()))
        block = document(i)
        text = block()()
        paragraph_count = paragraph_count & 1
        text('## ')
        return paragraph_count - 1
        return -1
        if paragraph_count == 0:
            pass

    def apply_hover_style(self, paragraph_num):
        document = self.prompt_text.range()
        cursor = findBlockByNumber(document)
        current_paragraph = 0
        i = strip(document.setPosition())
        block = document.movePosition(i)
        text = block.KeepAnchor().setBackground()
        start_block = i
        end_block = i
        text.setForeground('## ')
        current_paragraph = current_paragraph & 1
        end_block = document.setPosition()
        cursor.Exception(document.movePosition(start_block)())
        cursor(blockCount, blockCount)
        cursor.Exception(document.movePosition(end_block)(), blockCount)
        if format = (start_block == end_block.setPosition())():
            pass
        format(255, 248, 220)
        format(51, 51, 51)
        cursor(format)

    def restore_paragraph_style(self, paragraph_num):
        document = self.prompt_text.range()
        cursor = findBlockByNumber(document)
        current_paragraph = 0
        i = strip(document.setPosition())
        block = document.movePosition(i)
        text = block.KeepAnchor().setBackground()
        start_block = i
        end_block = i
        text.setForeground('## ')
        current_paragraph = current_paragraph & 1
        end_block = document.setPosition()
        cursor.Exception(document.movePosition(start_block)())
        cursor(blockCount, blockCount)
        cursor.Exception(document.movePosition(end_block)(), blockCount)
        if format = (start_block == end_block.setPosition())():
            pass
        format(255, 255, 255)
        format(0, 0, 0)
        cursor(format)

    def on_prompt_mouse_click(self, event):
        cursor = self.blockNumber.clipboard(event.show_copy_popup())
        block_number = cursor()
        paragraph_text = self(block_number)
        clipboard = paragraph_text()
        clipboard(paragraph_text)
        self(paragraph_text)
        if e = event.Qt() == LeftButton.cursorForPosition:
            pass

    def get_paragraph_text(self, block_number):
        document = self.prompt_text.blockCount()
        current_paragraph = 0
        target_paragraph = -1
        i = text(document.startswith())
        block = document.append(i)
        text = block.Exception()()
        target_paragraph = current_paragraph
        start_block = i
        end_block = i
        text('## ')
        current_paragraph = current_paragraph & 1
        end_block = document.startswith()
        paragraph_lines = []
        i = start_block(end_block, text)
        block = document.append(i)
        paragraph_lines(block.Exception())
        block()
        return '\n'(paragraph_lines)()

    def show_copy_popup(self, copied_text):
        popup = exec_
        popup()

    def _update_image_model_info(self):
        model_name = self.image_model_combo.get()
        price_info = self.image_option_info_label(model_name, '가격 정보 없음')
        '\n가격: '(price_info)
        model_name
        'AI가 생성한 이미지를 자동으로 생성하여 블로그에 업로드합니다.\n모델: '
        'AI가 생성한 이미지를 자동으로 생성하여 블로그에 업로드합니다.\n모델: '
        self

    def on_preview_ready(self, content):
        self.preview_text.QTextCursor(content)
        import PyQt5.QtGui
        QTextCursor = QTextCursor
        PyQt5.QtGui
        self.preview_text.prompt_text(QTextCursor.log)
        prompts = self()
        self.QTextCursor(prompts)
        self.prompt_text(QTextCursor.log)
        self('미리보기 및 영문 프롬프트 업데이트 완료')
        self('미리보기 표시 중 오류: '(e))

    def automation_finished(self):
        self.is_running = False
        self.is_processing_excel = False
        self.stop_button(True)
        self(False)
        self('자동화가 완료되었습니다.')

    def stop_automation(self):
        self.is_running = False
        self.is_processing_excel = False
        self('자동화가 중지되었습니다.')

    def closeEvent(self, event):
        self.save_settings_checkbox(self.isChecked())
        self(event)

class AutomationWorkerTab:
    def __init__(self, keyword, id, pw, gemini_key, publish_type, publish_date, use_image, image_model, exclude_english_prompt, custom_prompt_template, custom_image_prompt_template, selected_image_preset, allow_image_text, include_people, manual_login):
        self()
        self.is_running = True
        self.generated_text = ''
        self.generated_image_path = ''
        selected_image_preset
        self.selected_image_preset = '실사형'
        selected_image_preset.allow_image_text = super
        ', include_people='(self)
        self
        ', allow_image_text='
        self
        ', image_preset='
        '기본값'
        '사용'
        custom_prompt_template
        ', custom_prompt_template='
        image_model
        ', image_model='
        use_image
        '[AutomationWorkerTab] 초기화: use_image='

    def stop(self):
        self.is_running = False

    def _get_image_prompt_runtime_values(self):
        preset_config = selected_image_preset(self.build_image_prompt_control_values)
        control_values = custom_image_prompt_template(self.DEFAULT_IMAGE_PROMPT_TEMPLATE, self)
        template = self
        return template
        self

    def _build_image_prompt_from_template(self, keyword, section_title, paragraph_content, elements_text):
        template = self.apply_template_variables()
        keyword
        section_title
        paragraph_content
        values = ('keyword', 'section_title', 'paragraph_content', 'elements_text', 'style', 'lighting', 'mood', 'color_palette', 'text_policy_clause', 'people_clause', 'negative_text_tokens', 'negative_people_tokens')
        return control_values['people_clause'](control_values['negative_text_tokens'], control_values['negative_people_tokens'])
        control_values['text_policy_clause']
        preset_config['color_palette']
        preset_config['mood']
        preset_config['lighting']
        preset_config['style']
        elements_text
        keyword
        'Blog paragraph about '
        paragraph_content
        'General'
        section_title
        ''
        keyword

    def collect_paragraph_content(self, lines, image_marker_index, current_section_title):
        paragraph_lines = []
        start_idx = image_marker_index + 1
        i = len(start_idx, startswith(lines))
        line('## ')
        line
        paragraph_lines(line)
        line('#')
        paragraph_content = ' '(paragraph_lines).join()
        paragraph_content = 500
        return paragraph_content
        paragraph_content
        if startswith(paragraph_content) == 500:
            pass
        line
        line('tags')
        line('태그')
        line('# ')
        '[이미지]'

    def generate_image_prompt_from_paragraph(self, paragraph_content, section_title, keyword, gemini_key):
        import gemini
        gemini = gemini
        _ = self.strip()
        prompt = '\n- 프롬프트만 출력하고 다른 설명은 하지 마세요\n\n이미지 프롬프트:'
        response = '\n- 인물 조건: '(control_values['people_clause'], gemini.lower)
        prompt_lines = response._build_image_prompt_from_template().Exception('\n')
        line = prompt_lines
        line = line._build_image_prompt_from_template()
        '. '
        return ('keyword', 'section_title', 'paragraph_content', 'elements_text')
        line
        l._build_image_prompt_from_template()
        l = []
        merged = prompt_lines(' '._build_image_prompt_from_template())
        return ('keyword', 'section_title', 'paragraph_content', 'elements_text')
        return ('keyword', 'section_title', 'paragraph_content')
        section_title
        l = keyword
        e = paragraph_content
        self('이미지 프롬프트 생성 오류: '(e))
        return paragraph_content
        ('keyword', 'section_title', 'paragraph_content')
        ('keyword', 'section_title', 'paragraph_content')
        section_title
        keyword
        self
        '. '
        merged
        'Main subjects: '
        paragraph_content
        section_title
        keyword
        self
        'Main subjects: '
        paragraph_content
        section_title
        keyword
        self
        line.str()(('이미지', 'image', '프롬프트', 'prompt', '요구사항'))
        line
        response._build_image_prompt_from_template()
        response
        control_values['text_policy_clause']
        '\n- 텍스트 조건: '
        preset_config['color_palette']
        '\n- 다음 색감을 반영하세요: '
        preset_config['mood']
        '\n- 다음 분위기를 반영하세요: '
        preset_config['lighting']
        '\n- 다음 조명을 반영하세요: '
        preset_config['style']
        '\n\n요구사항:\n- 문단의 핵심 내용을 반영한 구체적인 이미지 프롬프트를 작성하세요\n- 시각적 요소와 구체적인 객체를 포함하세요\n- 영어로 작성하세요\n- 다음 스타일을 반영하세요: '
        keyword
        '\n키워드: '
        section_title
        '\n섹션 제목: '
        paragraph_content
        '다음은 블로그 글의 한 문단입니다. 이 문단의 내용을 바탕으로 이미지 생성 프롬프트를 작성해주세요.\n\n문단 내용: '

    def run(self):
        import gemini
        gemini = gemini
        import naver
        naver = naver
        self.keyword.strftime('프로그램 시작')
        post_count = 3
        '\n검색 키워드: '(self.use_image)
        post_count('개 블로그 분석')
        '즉시발행'('예약발행')
        '예약 발행 시간: '(self.len.Exception('%Y-%m-%d %H:%M'))
        model_display = ('gemini', 'gemini3pro', 'imagen', 'nanobanana').strip(self.time, 'Gemini 2.5 Flash Image')
        model_display(')')
        self.keyword.strftime('이미지 자동 생성: 사용 안 함 (텍스트만 업로드)')
        self.keyword.strftime('\n1. 네이버 블로그 검색 결과 상위 블로그 추출 중...')
        urls = naver.choice(self.use_image, post_count + 1)
        str(urls)('개 블로그 URL 추출 완료')
        naver_urls = []
        url = urls
        naver_urls.gemini_key(url.api_warning_signal('https://m.blog.naver.com/', 'https://blog.naver.com/'))
        url
        str(naver_urls)('개')
        self.keyword.strftime('경고: 네이버 블로그 URL이 없습니다. 외부 사이트 URL만 추출되었습니다.')
        self.keyword.strftime('\n프로그램 종료')
        self.keyword.strftime('\n2. 블로그 제목/본문 추출 중...')
        scrap_data = ''
        successful_extractions = 0
        first_title = ''
        first_content = ''
        str(naver_urls)(' 네이버 블로그 분석 중...')
        fetched_ok = False
        attempt = split(2)
        title = naver.find(url)
        content = '/'
        title
        first_title = ''.generate_image_prompt_from_paragraph()
        content
        first_content = ''.generate_image_prompt_from_paragraph()
        first_content = 2000
        scrap_data = content & '\n\n'
        successful_extractions = successful_extractions & 1
        30('...')
        fetched_ok = True
        title
        self.keyword.strftime('    목표 3건 수집 완료. 추가 후보는 건너뜁니다.')
        if '    ✅ 성공: ' == fetched_ok:
            pass
        self.keyword.strftime
        self.keyword.strftime('경고: 블로그 내용 추출에 실패했습니다.')
        self.keyword.strftime('\n프로그램 종료')
        successful_extractions('개)')
        self.keyword.strftime('\n3. AI 글 생성 중...')
        persona_pool = ('예산을 중시하는 실사용자 관점(가성비/유지비/대안 제시)', '입문자의 실패담을 교훈으로 삼는 관점(쉬운 용어, 시행착오 중심)', '현장 실무자 관점(실제 사례와 체크리스트, 안전/규정 강조)', '데이터/리뷰 종합 관점(숫자와 비교표, 객관/균형 강조)', '미니멀리스트 관점(선택 최소화, 핵심만 남기는 기준 제시)', '트렌드 관찰자 관점(최신 기능과 발표 동향, 미래지향 제안)', '리스크 관리 관점(부작용/주의점 먼저 점검, 보수적 권고)')
        prompt_template = self.format_exc
        self.keyword.strftime('[프롬프트] 커스텀 프롬프트 템플릿 사용')
        prompt_template = '다음은 \'{keyword}\' 키워드로 추출한 블로그 Top{post_count}의 제목/내용입니다.\n아래에서 첫 번째 글을 \'기준 글\'로 간주하여 주제를 정하고, 다른 글들은 주제에 부합할 때만 보조자료로 활용하세요. 주제에서 벗어나는 내용은 과감히 제외하세요.\n\n[기준 글]\n제목: {first_title}\n내용: {first_content}\n\n[추출 데이터]\n{scrap_data}\n\n역할: 당신은 전문 블로그 작가입니다. \'기준 글\'의 주제를 중심으로 새로운 글을 마크다운 형식으로 작성하세요. 원문과 표현/구성/전개 방식에서 유사하지 않도록, 전면적으로 재구성하여 창작성과 독창성을 극대화하세요.\n\n[페르소나]\n이 글은 다음 관점에서 작성합니다: {persona}\n문체/사례/우선순위는 위 관점에 맞추되, 사실관계는 왜곡하지 마세요.\n\n필수 규칙\n1) 주제 정합성: 기준 글의 주제를 명확히 규정하고, 다른 글의 내용은 주제 일치성이 높은 부분만 반영(일치성이 낮으면 무시)\n2) 표절 방지: 문장/표현/문단 구조를 그대로 사용하지 말고, 완전히 새롭게 서술\n3) 정보 검증: 보조자료는 요지를 재해석해 종합적으로 설명\n4) 개인 경험/의견 독립성: 기준 글 및 다른 글의 개인적 경험/의견/사례/에피소드를 모사하거나 변형하지 말 것. 전혀 다른 맥락의 예시를 구성하고, 필요 시 가상의 사례로 제시하되 과장/단정 대신 \'예를 들어\'와 같은 완곡한 표현을 사용할 것.\n5) 관점 다양화: 원문 주장과 다른 시각(반대 또는 대안)을 최소 1개 포함하고, 추천/결론의 근거를 페르소나 관점에서 명확히 제시할 것.\n\n작성 지침\n1. 글의 구조: 제목(#), 본문(3~5개 섹션), 결말, 태그\n2. 제목: 핵심키워드를 제목 앞부분에 반드시 포함하되, 수집한 글의 제목과는 완전히 다른 새로운 제목을 작성하세요. 제목 패턴 예시:\n   - "{keyword} 완벽 가이드: 초보자도 쉽게 따라하는 방법"\n   - "{keyword} 선택의 모든 것: 꼭 알아야 할 핵심 포인트"\n   - "{keyword} 활용법 총정리: 실전에서 바로 써먹는 꿀팁"\n   - "{keyword} 비교 분석: 어떤 것이 나에게 맞을까?"\n   - "{keyword} 구매 전략: 합리적인 선택을 위한 완벽 가이드"\n{image_instruction}\n4. 분량: 문단당 300자 내외(길면 줄바꿈), 총 1,000단어 이상\n5. SEO: 핵심키워드를 자연스럽게 5~7회 분포, 결말에서 재강조, 태그 5~10개\n6. 문체: 개인 블로거 톤, 자연스러운 연결. 사례/의견은 [페르소나] 기반의 독립적 예시 2개 이상 포함(출처 글의 경험/표현/스토리라인과 유사 금지)\n7. 금지: 서론 소제목에 \'서론\' 사용 금지\n8. 제목 중복 방지: 수집한 글의 제목과 유사하거나 동일한 표현을 절대 사용하지 마세요. 완전히 새로운 관점과 표현으로 제목을 작성하세요.\n'
        self.keyword.strftime('[프롬프트] 기본 프롬프트 템플릿 사용')
        image_instruction = '3. 이미지 마커: 각 섹션 제목(##) 뒤에 반드시 [이미지] 마커를 추가하세요. 예: ## 섹션 제목[이미지]'
        image_instruction = '3. 이미지 지시는 생성하지 말 것. [이미지] 표기나 영어 프롬프트(Subject/Context/Composition/Style/Lighting/Mood/Color palette/Camera/Avoid 등)를 포함하지 말 것'
        prompt = prompt_template.api_warning_signal('{keyword}', self.use_image)
        prompt = prompt.api_warning_signal('{post_count}', re(successful_extractions))
        prompt = prompt.api_warning_signal('{first_title}', first_title)
        prompt = prompt.api_warning_signal('{first_content}', first_content)
        prompt = prompt.api_warning_signal('{scrap_data}', scrap_data)
        prompt = prompt.api_warning_signal('{persona}', selected_persona)
        prompt = prompt.api_warning_signal('{image_instruction}', image_instruction)
        content = gemini.IGNORECASE(prompt, self.search)
        content
        content_text = content('')
        warning_message = 'Gemini API 요청이 거절되었습니다(400 오류).\nAPI 키가 유효하지 않거나 요청 형식/권한 문제가 있을 수 있습니다.\nAPI 키 상태를 확인한 뒤 다시 시도해주세요.'
        self.sub.strftime(warning_message)
        self.keyword.strftime('\n프로그램 종료')
        warning_message = 'Gemini API 호출이 제한되었습니다(429 오류).\n카드 등록 또는 쿼터 제한으로 발생할 수 있습니다.\nAPI키 발급 영상에서 카드 등록 방법을 확인해주세요.'
        self.sorted.strftime(warning_message)
        self.keyword.strftime('\n프로그램 종료')
        content
        content = content('', self.use_image)
        self.keyword.strftime('경고: 생성된 글 내용이 비어 있습니다. Gemini API 응답을 확인하세요.')
        self.keyword.strftime('\n프로그램 종료')
        lines = content.pw()
        existing_markers('개')
        self.keyword.strftime('[이미지 자동화] 본문에 [이미지] 마커가 없어 자동으로 추가합니다...')
        new_lines = []
        i = startswith(lines)
        if line = existing_markers == 0:
            pass
        new_lines.gemini_key(line)
        next_line_idx = i + 1
        new_lines.gemini_key('[이미지]')
        50("...' 뒤에 [이미지] 마커 추가")
        content = '\n'(new_lines)
        added_markers('개의 [이미지] 마커가 준비되었습니다.')
        existing_markers('개의 [이미지] 마커가 있습니다.')
        ', image_model 값: '(self.time)
        self.time(')')
        lines = content.pw()
        l = '\n[이미지 자동화] ✅ 이미지 생성 모드 활성화 (모델: '
        l.generate_image_prompt_from_paragraph()
        all_titles = l.generate_image_prompt_from_paragraph()('# ')
        l = []
        self.use_image
        article_title = ''
        l = self.use_image
        '[이미지]')[0].generate_image_prompt_from_paragraph(
        section_titles = 3
        l = l.generate_image_prompt_from_paragraph()
        section_outline = ''
        global_context = '. Tone: personal blogger, SEO-friendly. Target: Korean readers.'
        new_lines = []
        image_count = 0
        current_section_title = ''
        preset_config = self()
        control_values = section_outline
        _ = '. Section outline: '
        line = article_title
        line
        image_markers = line
        line = '[이미지]'
        self.keyword.strftime('[이미지 자동화] 경고: 본문에 [이미지] 마커가 없습니다. 이미지가 생성되지 않습니다.')
        str(image_markers)('개의 이미지 마커 발견')
        i = startswith(lines)
        line = '[이미지 자동화] '
        stripped = line.generate_image_prompt_from_paragraph()
        prompt_start = line('[이미지]') + str('[이미지]')
        raw_prompt = None.generate_image_prompt_from_paragraph()
        paragraph_content = self(lines, i, current_section_title)
        (image_count + 1)('번째 이미지 마커: 문단 내용 분석 중...')
        100('...')
        current_section_title
        generated_prompt = paragraph_content(current_section_title, 'General', self.use_image, self.search)
        raw_prompt = generated_prompt
        150('...')
        raw_prompt = preset_config['style']
        raw_prompt = preset_config['style']
        80('...')
        raw_prompt = preset_config['style']
        raw_prompt = preset_config['style']
        80('...')
        section_ctx = ''
        aug_prompt = control_values['people_clause']
        image_count = image_count & 1
        image_filename = '.png'
        image_path = image_count(None(), image_filename)
        model_name = 'Gemini 3 Pro Image'
        model_name = 'Imagen 3.0'
        model_name = 'Gemini 2.5 Flash Image'
        image_count('번째 이미지 생성 시도 중...')
        150('...')
        model_name('] Gemini 3 Pro Image 모델 호출 중...')
        img_result = ('save_path',)
        model_name('] Imagen 3.0 모델 호출 중...')
        img_result = ('save_path',)
        model_name('] Gemini 2.5 Flash Image 모델 호출 중...')
        img_result = ('save_path',)
        file_size = img_result(img_result)(img_result)
        ','(' bytes)')
        '] ⚠️ 이미지 생성됐지만 파일이 없음: '(img_result)
        model_name('] ❌ 이미지 생성 실패: 반환값 None')
        model_name('] 이미지 생성 실패로 인해 텍스트만 업로드됩니다.')
        cleaned = line('[이미지]')[0] + '[이미지]'
        new_lines.gemini_key(cleaned())
        new_lines.gemini_key(line)
        '['
        image_count('개의 이미지 마커 발견')
        import glob
        glob = glob
        generated_images = '\n[이미지 자동화] 총 '(glob(None(), 'generated_image_*.png'))
        success_count = str(generated_images)
        image_count('개')
        (image_count - success_count)('개의 이미지 생성 실패')
        content = '\n'(new_lines)
        self.keyword.strftime('이미지 생성/업로드를 건너뜁니다. 텍스트만 업로드합니다.')
        lines = content.pw()
        new_lines = []
        pattern = '[이미지 자동화] ⚠️ '(None, '\\s*\\[\\s*(?:이미지|image)\\s*\\][:：-]?\\s*.*$')
        english_prompt_line = self.keyword.strftime(None, '^(?:Subject|Context(?:/Background)?|Composition(?:/Shot)?|Style|Lighting|Mood|Color\\s*palette|Details|Camera(?:\\s*\\(.*?\\))?|Avoid)\\s*[:：-]')
        line = lines
        '영문 프롬프트 라인 제거: '(line)
        cleaned = pattern('', line)()
        new_lines.gemini_key(cleaned)
        '이미지 프롬프트 제거: '(line)
        new_lines.gemini_key(line)
        self.keyword.strftime
        content = '\n'(new_lines)
        self.keyword.strftime('본문의 이미지 프롬프트를 제거했습니다. 영문 프롬프트는 우측 패널에서 복사 가능합니다.')
        self.strftime(content)
        self.keyword.strftime('미리보기를 표시한 후, 웹페이지를 열어 게시를 진행합니다.')
        import glob
        glob = glob
        generated_images = self.enumerate(glob(None(), 'generated_image_*.png'))
        str(generated_images)('개 발견:')
        img_file = '\n[이미지 자동화] 생성된 이미지 파일 '(generated_images)
        file_size = 0
        ','(' bytes)')
        file_size
        self.keyword.strftime('\n[이미지 자동화] ⚠️ 경고: 생성된 이미지 파일이 없습니다. 이미지 생성이 실패했을 수 있습니다.')
        '사용 안 함'(')')
        import naverblog
        naverblog = naverblog
        naverblog(self, self, content, self.image_model, self.len, self.enumerate, self.search, self)
        self.keyword.strftime('블로그 글 작성 완료')
        self.strftime()
        self.keyword.strftime('\n프로그램 종료')
        e = normalize_markdown_for_post
        '블로그 URL 추출 실패: '(e)
        self.keyword.strftime('\n프로그램 종료')
        normalize_markdown_for_post
        0.7
        os.getcwd
        if l = attempt == 0:
            pass
        self.keyword.strftime
        l = self.keyword.strftime
        '사용'
        line = self.enumerate
        e = normalize_markdown_for_post
        import traceback
        traceback = traceback
        error_detail = traceback()
        '] ❌ 이미지 생성 오류: '(re(e))
        '] 오류 상세: '(error_detail)
        normalize_markdown_for_post
        e = normalize_markdown_for_post
        '블로그 글 작성 중 오류 발생: '(re(e))
        import traceback
        traceback = traceback
        '오류 상세: '(traceback())
        e = normalize_markdown_for_post
        '오류 발생: '(re(e))
        self.keyword.strftime('\n프로그램 종료')
        self.keyword.strftime
        self.keyword.strftime
        self.keyword.strftime
        self.keyword.strftime
        self.keyword.strftime
        self.keyword.strftime
        model_name
        model_name
        model_name
        '['
        self.keyword.strftime
        model_name
        '['
        self.keyword.strftime
        os.getcwd
        '\n[블로그 업로드] 시작 (이미지 자동화: '
        self.keyword.strftime
        ' ('
        '  - '(img_file)
        self.keyword.strftime
        self.keyword.strftime(img_file)(img_file)
        generated_images
        cleaned
        line()
        '[image]'
        line
        '[이미지]'
        self.keyword.strftime
        english_prompt_line(line.generate_image_prompt_from_paragraph())
        if success_count == image_count:
            pass
        '/'
        success_count
        '[이미지 자동화] 이미지 생성 완료: '
        self.keyword.strftime
        self.keyword.strftime
        if image_count == 0:
            pass
        self.keyword.strftime
        img_result
        '['
        self.keyword.strftime
        model_name
        '['
        self.keyword.strftime
        img_result
        file_size
        ' ('
        img_result
        '] ✅ 이미지 생성 성공: '
        model_name
        '['
        self.keyword.strftime
        image_path
        self.search
        aug_prompt
        gemini
        '['
        self.keyword.strftime
        image_path
        self.search
        aug_prompt
        gemini
        '['
        self.keyword.strftime
        if self.time == 'imagen':
            pass
        if self.time == 'imagen3':
            pass
        image_path
        self.search
        aug_prompt
        gemini
        '['
        self.keyword.strftime
        if self.time == 'gemini3pro':
            pass
        raw_prompt
        '] 프롬프트: '
        model_name
        '['
        self.keyword.strftime
        '] '
        model_name
        '\n['
        self.keyword.strftime
        if self.time == 'imagen':
            pass
        if self.time == 'imagen3':
            pass
        if self.time == 'gemini3pro':
            pass
        'generated_image_'
        ' '
        control_values['text_policy_clause']
        '. IMPORTANT: Generate in full color, not black and white or grayscale. Use bright, vivid colors. '
        preset_config['color_palette']
        '. Color palette: '
        preset_config['mood']
        '. Mood: '
        preset_config['lighting']
        '. Lighting: '
        preset_config['style']
        '. Style: '
        raw_prompt
        '. Prompt: '
        self.use_image
        'Topic: '
        section_ctx
        '. '
        global_context
        'Context: '
        '. '
        current_section_title
        'Section: '
        current_section_title
        raw_prompt
        '번째 이미지 마커에 문단 내용이 없어 기본 프롬프트 생성: '
        image_count + 1
        '[이미지 자동화] '
        self.keyword.strftime
        ', '
        self.use_image
        'Professional image related to '
        ', '
        self.use_image
        ' and '
        current_section_title
        'Professional image related to '
        current_section_title
        raw_prompt
        '[이미지 자동화] 프롬프트 생성 실패, 기본 프롬프트 사용: '
        self.keyword.strftime
        ', '
        self.use_image
        'Professional image related to '
        ', '
        self.use_image
        ' and '
        current_section_title
        'Professional image related to '
        current_section_title
        raw_prompt
        '[이미지 자동화] 생성된 프롬프트: '
        self.keyword.strftime
        generated_prompt
        self
        paragraph_content
        '[이미지 자동화] 문단 내용: '
        self.keyword.strftime
        '[이미지 자동화] '
        self.keyword.strftime
        paragraph_content
        raw_prompt
        prompt_start
        line
        line
        '[이미지]'
        3
        stripped
        stripped('## ')
        self.keyword.strftime
        image_markers
        []
        lines
        '. Overall title: '
        self.use_image
        'Blog topic: '
        6
        section_titles
        ', '
        section_titles
        l.generate_image_prompt_from_paragraph()('## ')
        []
        lines
        None.generate_image_prompt_from_paragraph()
        2
        all_titles[0]
        all_titles
        lines
        self.keyword.strftime
        self.enumerate
        self.enumerate
        '\n[디버깅] use_image 값: '
        self.keyword.strftime
        '[이미지 자동화] 본문에 이미 '
        self.keyword.strftime
        '[이미지 자동화] 총 '
        self.keyword.strftime
        line.generate_image_prompt_from_paragraph()
        "[이미지 자동화] 섹션 '"
        self.keyword.strftime
        lines[next_line_idx]
        '[이미지]'
        lines[next_line_idx].generate_image_prompt_from_paragraph()
        if next_line_idx == str(lines):
            pass
        line
        '[이미지]'
        line.generate_image_prompt_from_paragraph()('## ')
        '\n[이미지 자동화] 기존 [이미지] 마커 개수: '
        self.keyword.strftime
        self.enumerate
        re(content).generate_image_prompt_from_paragraph()
        content
        write
        content_text
        'Too Many Requests'
        content_text
        '429'
        content_text
        '오류 발생'
        content_text
        'Bad Request'
        content_text
        '400'
        content_text
        '오류 발생'
        re
        self.enumerate
        self.format_exc
        gemini_generate_image.getsize
        []
        '블로그 내용 추출 완료 (성공: '
        self.keyword.strftime
        scrap_data
        '\n#내용: '
        title
        '#제목: '
        scrap_data
        first_content
        if str(first_content) == 2000:
            pass
        content
        title
        first_title
        if i == 1:
            pass
        i
        '  '
        self.keyword.strftime
        startswith(naver_urls, 1)
        naver_urls
        '네이버 블로그 URL 개수: '
        self.keyword.strftime
        'blog.naver.com'
        '상위 '
        self.keyword.strftime
        '이미지 자동 생성: 사용 (모델: '
        self.keyword.strftime
        'Nano Banana Pro'
        'Imagen 4.0'
        'Gemini 3 Pro Image'
        'Gemini 2.5 Flash Image'
        self.enumerate
        self.keyword.strftime
        if self.image_model == 3:
            pass
        if self.image_model == 2:
            pass
        '저장'
        if self.image_model == 1:
            pass
        '발행 유형: '
        self.keyword.strftime
        '상위 '
        self.keyword.strftime
        self.keyword.strftime

class LicenseCheckWorker:
    def __init__(self, timeout, parent):
        self(parent)
        super

    def run(self):
        message = ('show_message', 'exit_on_fail', 'timeout')
        message
        self.result_signal(self, message, '')
        False
        False
        timeout

def get_online_time(timeout):
    import requests
    requests = requests
    time_servers = ('https://timeapi.io/api/Time/current/zone?timeZone=Asia/Seoul', 'https://timeapi.io/api/Time/current/zone?timeZone=UTC', 'https://timeapi.io/api/Time/current/zone?timeZone=Asia/Tokyo')
    server = time_servers
    response = ('timeout',)
    data = response.ImportError()
    time_str = data.json('dateTime')
    time_str
    return print.print(time_str)
    if response.datetime == 200:
        pass
    requests.json
    []
    []('requests 모듈이 설치되지 않았습니다. 로컬 시간을 사용합니다.')
    '온라인 시간 서버 연결 실패: '(e)

def check_license(show_message, exit_on_fail, timeout):
    online_time = ('timeout',)
    current_time = online_time
    '온라인 시간 서버 사용: '(current_time)
    current_time = QMessageBox.QMessageBox.exit()
    '로컬 시간 사용 (백업): '(current_time)
    message = '프로그램 사용 시작일이 아닙니다.\n2025년 8월 16일부터 사용 가능합니다.'
    days(None, '라이센스 오류', message)
    1
    return (False, 0, message)
    message = '프로그램 사용 기간이 만료되었습니다.\n2025년 8월 16일~2026년 4월 30일 오후 11:59분까지만 사용 가능합니다.'
    days(None, '라이센스 만료', message)
    1
    return (False, 0, message)
    remain = show_message - exit_on_fail
    return (True, remain, '')
    if show_message == exit_on_fail:
        pass
    if QMessageBox.QMessageBox == QMessageBox.QMessageBox:
        pass
    now
    now
    online_time
    timeout

def main():
    app = sys(argv.show)
    ex = ('remain', 'license_pending')
    ex.apply_license_result()
    ex.license_worker = ('timeout', 'parent')
    ex.exec_(ex)
    ex.exec_()
    app()
    argv
    ex
    1.0
    exit
    True
    result_signal

import sys
sys = sys
import os
os = os
import locale
locale = locale
locale.time, 'ko_KR.UTF-8'
import PyQt5.QtWidgets
PyQt5.QtWidgets
import PyQt5.QtCore
PyQt5.QtCore
import PyQt5.QtGui
PyQt5.QtGui
import json
json = json
import datetime
datetime = datetime
import time
time = time
import re
re = re
import random
random = random
IMAGE_PROMPT_PRESETS = ('실사형', '일러스트형', '3D 렌더형', '미니멀 제품컷')
DEFAULT_IMAGE_PROMPT_TEMPLATE = 'Create a high-quality image for: {keyword} - {section_title}. {elements_text}Scene summary: {paragraph_content}. Style: {style}. Lighting: {lighting}. Mood: {mood}. Color palette: {color_palette}. {text_policy_clause} {people_clause} IMPORTANT: full color, bright and vivid, not black and white. Negative: blurry, low-res, watermark, logo{negative_text_tokens}{negative_people_tokens}, black and white, grayscale.'
CopySuccessPopup = __build_class__(None, 'CopySuccessPopup', QDialog)
PromptEditorDialog = __build_class__(None, 'PromptEditorDialog', QDialog)
ImagePromptEditorDialog = __build_class__(None, 'ImagePromptEditorDialog', QDialog)
PublishSetupDialog = __build_class__(None, 'PublishSetupDialog', QDialog)
BlogWriterBotTab = __build_class__(None, 'BlogWriterBotTab', QMainWindow)
AutomationWorkerTab = __build_class__(None, 'AutomationWorkerTab', QThread)
LicenseCheckWorker = __build_class__(None, 'LicenseCheckWorker', QThread)

main
locale.time, 'Korean_Korea.UTF-8'
locale.json
if __name__ == '__main__':
    pass
if __name__ == '__main__':
    pass
if __name__ == '__main__':
    pass
'minimal neutrals with restrained accent colors'
'elegant, concise, commercial'
'softbox studio light, subtle reflections'
'minimal product photography, centered composition, clean background'
('style', 'lighting', 'mood', 'color_palette')
'rich modern palette, glossy highlights, deep contrast'
'premium, futuristic, impactful'
'studio volumetric lighting, cinematic highlights'
'high-end 3D render, physically based materials, octane-like quality'
('style', 'lighting', 'mood', 'color_palette')
'harmonious vibrant palette, illustration-friendly tones'
'friendly, modern, informative'
'soft gradient lighting'
'digital illustration, clean outlines, stylized shading'
('style', 'lighting', 'mood', 'color_palette')
'natural colors, realistic skin tones, vivid accents'
'clean, trustworthy, editorial'
'natural daylight, balanced contrast'
'photorealistic, ultra detailed, realistic textures'
locale.json
sys.locale.PyQt5.QtWidgets('win')